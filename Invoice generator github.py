import os
import pandas as pd
import numpy as np
import re
import ipywidgets as widgets
from IPython.display import display, clear_output
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, numbers, PatternFill, Color
from openpyxl.worksheet.page import PageMargins
from openpyxl.drawing.image import Image
import math


#############################################################################
######################## --- GitHub VERSION --- #############################
#############################################################################

## 1. Th GitHub version is essentially the same version that i actually do use, the minor difference being that, the data which is being used in the code is changed slightly, to maintain confidentiality of the actuall data. 

## 2. Most of the functions in this code are in def function form, becasue of the 'ipywidgets' usage, otherwise there wouldn't be need to use so much def functions ###

#===============================================================================================================================================#


############################################################
#############################################################
###                   CODE STRUCTURE                     ####

# 1. Spreadsheet prep -- For extracting info from excel and dividing them into different dataframes for each sheet.
# 2. Dataframes -- Each spreadsheet dataframe is divided into its own product category group.
# 3. Generator -- Deciphers the dataframes into different elements that allows to create random item quantity and total sums based on inputs.
# 4. Filtering -- Regex filters for the product dataframe groups.
# 5. Interface -- Main interface product group checkboxes and product to category grouping. Checks main product category checkboxes.
# 6. Filter selection -- Filter checkbox and assignment.
# 7. Common items and sliders -- Creates common items checkbox and adjusts sliders automaticly based on filters and if there are on, or not.
# 8. Filter updaters -- Creates signals for observers to assign custom filter checkboxes based on assgined Checkbox_dictionary_Filter.
# 9. Dataframe filtering -- Creates filter VBoxes for storage and creates filtered_df, by applying selected custom filters into a single mask.
# 10. Max sliders and filter positions -- Creates number of sliders and assigns the possitions of filters on UI (left and right).
# 11. Invoice settings -- Controls the price generation regarding upper and lower price limit via Invoice_generator.
# 12. Client info -- Controls invoice client info inputs and stores client info.
# 13. Invoice excel formating - Controls the excel invoice formating.
# 14. Display -- Main outputs of widgets for the UI.



#############################################################
##############################################################
###                  HOW THE CODE WORKS                   ####

# 1. From the given path, it takes already preprepared excel file with all spreadsheets and seperates it all into different dataframes, based on what columns needs to be extracted and changing the name of those columns to english ones.
# 2. Dataframes are created using regex filtering for each product group based on index sepperation. The product groups are selected from the general category that is actually usefull or at least is used on common basis. Then those product groups are assigned via dictionary for access.
# 3. The generator part is responsible for invoice structure generation, based on column inputs.

    # Probability(num) - is just a simple random probability generator, based on the len(num) input, if the input is 5, it will generatre random       probability -p- for each 5 inputs.
    # Product_price_generator(args) - depending on the filter mode that is selected, 'normal' or 'common', first part of the function is               dedicated for probability creation, based on 'normal' or 'common' mode, which are later used for price calculation. when it comes to             common mode, different maps can be used and weight assginment if needed. Once the probabilites, higher and lower price limits are                calculated based on the manual inputs in the UI. These inputs allow to create a price range for which the final price of the invoice             will be generated based on the amount of products are selected. The amount of products that are selected, dirrectly correlate to the             probabilites, that are also used calculate Quantity variable. This varriable helps to calculate total sum - 'Final_sum' variable for each        row of products that are selected.
    # Disassembler(dataframe) - simple function that transform dataframe into a dictionary, while also adding 'Index' column.
    # Invoice_assembler() and simple_Invoice_assembler() - Main invoice data frame assmebler, where it creates basically all the invoice               strucutre that is in the dataframe format. There are 2 different assembler fucntions becasue 'Invoice_generator' later decides which one         to used, based on what filter is selected, 'common items' or just 'normal' modes. If the common items are selected, Invoice_assembler will       be use, otherwise simple_Invoice_assembler will be used, once normal mode is selected.

# 4. Just basic regex filters that are later used to filter out thorugh different assigned groups.
# 5. Interface part is responsible for creating main category checkboxes that are observed through Update_items function. It allows further access for deeper filter selection through the collection of observers at the end of the function. This part also stores outputs for the  updaters that are directly related to this function, which observers are observering. Meaning that if one category is selected via Checkbox, it stores the info in Item_checkboxes that allows for other functions to work like updaters, that further branch out into specific filters.
# 6. In Filter selection part, seperate Checkboxes are being created to store regix filters that are later being assigned in a dictionary which each corresponding regex filter varriable. Once each filter is assigned, then a the main dictionary - 'Checkbox_dictionary_Filter' is created, that stores information on which dataframes is compatable with which filter, allowing custom filter -manual- assignment, depending if the actuall excel data has data for the propper filter selection. This section also stores custom filter dictionaries that allows to branch off into multilayer filtering, e.g. pipe type.
# 7. Next, the custom slider settings are created based on the selected dataframe from the main category. The main takeaway from this section is the first lines of the 'update_max_parts_slider' function, specificly -hasattr(Invoice_filtering, 'last_filtered_df')- part. This part reacts to the last_filtered_df atribute that is created in 'Invoice_filtering' function. Meaning that in that function, it stores custom selected filters in a mask that is later assigned to the selected_df, that is stored with the 'last_filtered_df' attribute. Meaning it creates Bool outcome of True/False. This is important becasuse for the initial 'update_max_parts_slider' function, it determines what is the custom slider max selection. Meaning that if filters are selected, it will change according automaticly, if not, it will basically set to the default selected category max product count (if product group contains 12 rows e.g. the max will be 12, if with filters selected number reduces to 4, max will be 4).  
# 8. Stores updaters that allows for further filter selection, once the main categories and sub-groups were selected from the Item_checkboxes. It basically stores updaters that refrence to the Checkbox_dictionary_Filter for the custom filter assignment. Stores the signals that are catched by the observers in the Update_items function, that allows sub-groups to appear and custom filters to be shown from the custom filter boxes.
# 9. Once the updaters are created, next VBoxes are created to store the selected filter checkboxes, while 'Invoice_filtering' function applies the selected filters to the dataframe and returns filtered_df, that can be affected my normal filters or 'common' filter method and its sliders.
# 10. This part is also responsible mainly for custom slider activity. Basically 'update_group_common_filters' also works in tender with custom sliders, even thou it is put later down the code, as it reacts to what groups are selected and reacts to the amount of groups are selected, allowing to match the slider amount. This section also controls slider settings and how many can there be. Lastly, in this part, filter palcement is controler via HBoxes and VBoxes, assigning them to the normal filters on the left side, and 'common' item filters and its sliders to the 'special filters' side that is on the right.
# 11. 'Invoice_generator' is in essence, responsible for price input and calculation. Here the text boxes are stored which can receive gaol amount, upper and lower limtis of the amount. and 'Invoice_generator' raects to the inputs accordingly, in regarnds to which specific products groups it is assigned. Meaning that only the specified groups of dataframes will react to the inputs of these custom price settings.
# 12. This part is responsible for client info in the invoice. It has multiple functions. It has a dropdown menu, that allows to select predetermined 'common' clients, that stores all the necessary info like, company name, location, company's code and VAT code. Here you can also manually add custom clients for future preference. If a predetermined client is selected, all the info is already filled in, otherwise 'Empty' option is available, that allows you to custom input the details.
# 13. Next up is the excel formating and invoice sheet assembling using 'openpyxl' library. This section also contains word to text converters, that allows to convert final invoice amount into word format that you woudn't need to do it manually. Also this part contains reactive part to the finised 'Invoice' dataframe that is created in 'Invoice_generator' function, that allows to extend invoice cells automaticly, in regards to how many rows there are in the dataframe. Lastly it has the pathway to the invoice export that reacts to the 'File_printer' and print invoice button, alongside the file naming system, that takes into consideration most possible errors.
# 14. Lastly the outputs that ties everything together, helps to make the UI and contains main UI buttons and sums up all the functions of the code, for the final outputs.

##############################################################
#############################################################


#===============================================================================================================================================#
# 1. Spreadsheet prep:


##################################
######## !!! ATENTION !!! ########
##################################

# Regarding the paths, please fill in your saving path, which can be desktop, and the excel Database file path.
### Stock:
Path = "C:/Users/kmito/Desktop/Github test data/Price Database Github.xlsx"
Saving_path = "C:/Users/kmito/Desktop/"
Logo_path = "C:/Users/kmito/Desktop/Github test data/Logo.png"

# By changing the RGB numbers, you can adjust the UI color, since right now it is set to default, so it would match the light background, blue buttons color.
UI_color = "65, 128, 201" # < -- Default Blue UI 



Name_list = ['Kodas', 'Pavadinimas', 'Tiekėjas','Tiekėjo kaina Eur.', 'Mano', 'Rinkos', 'Pelnas Eur.', 'Mano kaina Be PVM', 'PVM',\
             'Galutinė kaina']
Fixed_names = ['Code', "Name", 'Dealer', "Dealer_price", "My_proc", "Market_proc", "Profit", "Price_no_tax", "Tax", "Final_price"]
Column_name_map = dict(zip(Name_list, Fixed_names))

Stock_df = pd.read_excel(Path, header=10, sheet_name="Inventorius", dtype={"Kodas":str}).iloc[:, :10]
Stock_alt_price = pd.read_excel(Path, header=10, sheet_name="Inventorius", dtype={"Kodas":str}).iloc[:, [14]]
Stock_df.iloc[:, 9] = Stock_alt_price.values
Stock_df = Stock_df.dropna()
Filtered_stock = Stock_df.copy()


Filtered_stock = Filtered_stock.rename(columns=Column_name_map)
Filtered_stock = Filtered_stock.set_index("Code")

Filtered_stock["My_proc"] = (Filtered_stock["My_proc"] * 100).astype("Int64")
Filtered_stock["Market_proc"] = (Filtered_stock["Market_proc"] * 100).astype("Int64")
Filtered_df = Filtered_stock.drop(columns="Dealer") 


### Plastic:
Plastic_df = pd.read_excel(Path, header=10, sheet_name="Plastikas", dtype={"Kodas":str}).iloc[:, :10]
Plastic_alt_price = pd.read_excel(Path, header=10, sheet_name="Plastikas", dtype={"Kodas":str}).iloc[:, [14]]
Plastic_df.iloc[:, 9] = Plastic_alt_price.iloc[:, 0].values
Plastic_df = Plastic_df.dropna().rename(columns=Column_name_map)
Plastic_df = Plastic_df.set_index("Code")

Plastic_df["My_proc"] = (Plastic_df["My_proc"] * 100).round(1).astype("Int64")
Plastic_df["Market_proc"] = (Plastic_df["Market_proc"] * 100).round(1).astype("Int64")
Plastic_df = Plastic_df.drop(columns="Dealer")


### Bronze:
Bronze_df = pd.read_excel(Path, header=10, sheet_name="Bronza", dtype={"Kodas":str}).iloc[:, :10]
Bronze_alt_price = pd.read_excel(Path, header=10, sheet_name="Bronza", dtype={"Kodas":str}).iloc[:, [14]] 
Bronze_df.iloc[:, 9] = Bronze_alt_price.iloc[:, 0].values
Bronze_df = Bronze_df.dropna().rename(columns=Column_name_map)
Bronze_df = Bronze_df.set_index("Code")

Bronze_df["My_proc"] = (Bronze_df["My_proc"] * 100).round(1).astype("Int64")
Bronze_df["Market_proc"] = (Bronze_df["Market_proc"] * 100).round(1).astype("Int64")
Bronze_df = Bronze_df.drop(columns="Dealer")


### Water:
Water_df = pd.read_excel(Path, header=10, sheet_name="Vonios kambarys", dtype={"Kodas": str}).iloc[:, :10]
Water_alt_price = pd.read_excel(Path, header=10, sheet_name="Vonios kambarys", dtype={"Kodas": str}).iloc[:, [14]]
Water_df.iloc[:, 9] = Water_alt_price.iloc[:, 0].values
Water_df = Water_df.dropna().rename(columns=Column_name_map)
Water_df = Water_df.set_index("Code")

Water_df["My_proc"] = (Water_df["My_proc"] * 100).round(1).astype("Int64")
Water_df["Market_proc"] = (Water_df["Market_proc"] * 100).round(1).astype("Int64")
Water_df = Water_df.drop(columns="Dealer")


### Sewage:
Sewage_df = pd.read_excel(Path, header=10, sheet_name="Kanalizacija", dtype={"Kodas":str}).iloc[:, :10]
Sewage_alt_price = pd.read_excel(Path, header=10, sheet_name="Kanalizacija", dtype={"Kodas": str}).iloc[:, [14]]
Sewage_df.iloc[:, 9] = Sewage_alt_price.iloc[:, 0].values
Sewage_df = Sewage_df.dropna().rename(columns=Column_name_map)
Sewage_df = Sewage_df.set_index("Code")

Sewage_df["My_proc"] = (Sewage_df["My_proc"] * 100).round(1).astype("Int64")
Sewage_df["Market_proc"] = (Sewage_df["Market_proc"] * 100).round(1).astype("Int64")
Sewage_df = Sewage_df.drop(columns="Dealer")


### Chrome:
Chrome_df = pd.read_excel(Path, header=10, sheet_name="Chrom. detalės", dtype={"Kodas": str}).iloc[:, :10]
Chorme_alt_price = pd.read_excel(Path, header=10, sheet_name="Chrom. detalės", dtype={"Kodas": str}).iloc[:, [14]]
Chrome_df.iloc[:, 9] = Chorme_alt_price.iloc[:, 0].values
Chrome_df = Chrome_df.dropna().rename(columns=Column_name_map)
Chrome_df = Chrome_df.set_index("Code")

Chrome_df["My_proc"] = (Chrome_df["My_proc"] * 100).round(1).astype("Int64")
Chrome_df["Market_proc"] = (Chrome_df["Market_proc"] * 100).round(1).astype("Int64")
Chrome_df = Chrome_df.drop(columns="Dealer")


### Heating:
Heating_df = pd.read_excel(Path, header=10, sheet_name="Šildymas", dtype={"Kodas": str}).iloc[:, :10]
Heating_alt_price = pd.read_excel(Path, header=10, sheet_name="Šildymas", dtype={"Kodas": str}).iloc[:, [14]]
Heating_df.iloc[:, 9] = Heating_alt_price.iloc[:, 0].values
Heating_df = Heating_df.dropna().rename(columns=Column_name_map)
Heating_df = Heating_df.set_index("Code")

Heating_df["My_proc"] = (Heating_df["My_proc"] * 100).round(1).astype("Int64")
Heating_df["Market_proc"] = (Heating_df["Market_proc"] * 100).round(1).astype("Int64")
Heating_df = Heating_df.drop(columns="Dealer")


### Metal:
Metal_df = pd.read_excel(Path, header=10, sheet_name="Metalas", dtype={"Kodas": str}).iloc[:, :10]
Metal_alt_price = pd.read_excel(Path, header=10, sheet_name="Metalas", dtype={"Kodas": str}).iloc[:, [14]]
Metal_df.iloc[:, 9] = Metal_alt_price.iloc[:, 0].values
Metal_df = Metal_df.dropna().rename(columns=Column_name_map)
Metal_df = Metal_df.set_index("Code")

Metal_df["My_proc"] = (Metal_df["My_proc"] * 100).round(1).astype("Int64")
Metal_df["Market_proc"] = (Metal_df["Market_proc"] * 100).round(1).astype("Int64")
Metal_df = Metal_df.drop(columns="Dealer")



## Available ##

# -- Filtered_df - Inventory
# -- Plastic_df - Plastic
# -- Bronze_df - Bronze
# -- Water_df - Bathroom parts (Water lines)
# -- Sewage_df - Sewage
# -- Chrome_df - Chrome parts
# -- Heating_df - Parts related to heating 
# -- Metal_df - Metal parts


#===============================================================================================================================================#
# 2. Dataframes:

### Chromas:
Chrome_parts = Chrome_df[Chrome_df.index.str.contains(r'^005(?:0[0-9]|[1-4][0-9])$')]
Chrome_extenders = Chrome_df[Chrome_df.index.str.contains(r'^005(?:[5-9][0-9])$')]


### Kanalizacija:
Sewage_pipes = Sewage_df[Sewage_df.index.str.contains(r'^006\d{2}$')]
Sewage_pumps = Sewage_df[Sewage_df.index.str.contains(r'^007\d{2}$')]
Sewage_pipe_clams = Sewage_df[Sewage_df.index.str.contains(r'^008\d{2}$')]
Sewage_chemicals = Sewage_df[Sewage_df.index.str.contains(r'^009\d{2}$')]
Sewage_siphon = Sewage_df[Sewage_df.index.str.contains(r'^010(?:0[0-9]|[1-4][0-9])$')]
Sewage_Toilet_connect = Sewage_df[Sewage_df.index.str.contains(r'^010(?:[5-9][0-9])$')]


### Water:
Hoses_bolts = Water_df[Water_df.index.str.contains(r'^W01\d{2}$')]
Toilets = Water_df[Water_df.index.str.contains(r'^(?:W00(?:0[0-9]|[1-9][0-9])|011\d{2})$')]
Toilet_parts_water = Water_df[Water_df.index.str.contains(r'^012\d{2}$')]
Consumables = Water_df[Water_df.index.str.contains(r'^013\d{2}$')]
Faucets = Water_df[Water_df.index.str.contains(r'^V00\d{2}$')]
Sinks = Water_df[Water_df.index.str.contains(r'^P00\d{2}$')]
Shower_sets = Water_df[Water_df.index.str.contains(r'^D00\d{2}$')]
Valves_water = Water_df[Water_df.index.str.contains(r'^VT\d{4}$')]
Rubber_seals_water = Water_df[Water_df.index.str.contains(r'^T0\d{3}$')]
Filters_meters = Water_df[Water_df.index.str.contains(r'^(?:014(?:0[0-9]|[1-4][0-9]))$')]
Miscellaneous_water = Water_df[Water_df.index.str.contains(r'^(?:014(?:[5-9][0-9]))$')]
Pipe_clams_water = Water_df[Water_df.index.str.contains(r'^015\d{2}$')]
Water_heaters = Water_df[Water_df.index.str.contains(r'^WH\d{3}$')]
Press_fittings = Water_df[Water_df.index.str.contains(r'^01[6-7][0-9]{2}$')]
Press_pipes = Water_df[Water_df.index.str.contains(r'^(?:018[0-4][0-9])$')]
Screws = Water_df[Water_df.index.str.contains(r'^(?:018[5-9][0-9])$')]
PE_pipe = Water_df[Water_df.index.str.contains(r'^(?:019\d{2})$')]


### Heating:
Radiators = Heating_df[Heating_df.index.str.contains(r'^R0\d{3}$')]
Heating_valves = Heating_df[Heating_df.index.str.contains(r'^(?:020\d{2})$')]
Thermo_manometers = Heating_df[Heating_df.index.str.contains(r'^(?:021(?:[0-4][0-9]))$')]
Heat_consumables = Heating_df[Heating_df.index.str.contains(r'^(?:021(?:[5-9][0-9]))$')]
Fiberglass_insulation = Heating_df[Heating_df.index.str.contains(r'^(?:022\d{2})$')]
Expansion_tanks = Heating_df[Heating_df.index.str.contains(r'^IND\d{3}$')]
Exp_tank_mounts = Heating_df[Heating_df.index.str.contains(r'^L\d{4}$')]
Manifold_parts = Heating_df[Heating_df.index.str.contains(r'^(?:024(?:[0-4][0-9]))$')]
Heating_manifold = Heating_df[Heating_df.index.str.contains(r'^(?:024(?:[5-9][0-9]))$')]
Gas_water_heater = Heating_df[Heating_df.index.str.contains(r'^(?:025(?:[0-4][0-9]))$')]
Gas_heater_parts = Heating_df[Heating_df.index.str.contains(r'^(?:025(?:[5-9][0-9]))$')]
Gas_fittings = Heating_df[Heating_df.index.str.contains(r'^DJ\d{3}$')]
Coolant = Heating_df[Heating_df.index.str.contains(r'^(?:026(?:[0-4][0-9]))$')]

### Plastic:
Fittings_plastic = Plastic_df[Plastic_df.index.str.contains(r'^030(?:[0-4][0-9])$')]
Reductions_plastic = Plastic_df[Plastic_df.index.str.contains(r'^030(?:[5-9][0-9])$')]
Modular_plastic = Plastic_df[Plastic_df.index.str.contains(r'^031\d{2}$')] 
Pipes_plastic = Plastic_df[Plastic_df.index.str.contains(r'^032\d{2}$')]
Polyethylene_isolation = Plastic_df[Plastic_df.index.str.contains(r'^033(?:[0-4][0-9])$')]
PVC_pipes = Plastic_df[Plastic_df.index.str.contains(r'^034(?:[5-9][0-9])$')]
PVC_fittings = Plastic_df[Plastic_df.index.str.contains(r'^035\d{2}$')]
PE_fittings = Plastic_df[Plastic_df.index.str.contains(r'^036\d{2}$')]


### Bronze:
Fittings_bronze = Bronze_df[Bronze_df.index.str.contains(r'^040\d{2}$')]


### Metal:
Black_metal = Metal_df[Metal_df.index.str.contains(r'^045\d{2}$')]
Stainless_steel = Metal_df[Metal_df.index.str.contains(r'^046\d{2}$')]
Flanges = Metal_df[Metal_df.index.str.contains(r'^047\d{2}$')]
Zinc_fittings = Metal_df[Metal_df.index.str.contains(r'^048\d{2}$')]


### Stock:
Plastic_stock = Filtered_df[Filtered_df.index.str.contains(r'^051\d{2}$')]
Modular_plastic_stock = Filtered_df[Filtered_df.index.str.contains(r'^052\d{2}$')]
Brackets_stock = Filtered_df[Filtered_df.index.str.contains(r'^053\d{2}$')]
Other_stock = Filtered_df[Filtered_df.index.str.contains(r'^056\d{2}$')]



Combined_df = {"Chrome parts": Chrome_parts,
               "Extenders": Chrome_extenders,
               "Sewage pipes": Sewage_pipes,
               "Pumps": Sewage_pumps,
               "Sewage clams": Sewage_pipe_clams,
               "Chemicals": Sewage_chemicals,
               "Siphons": Sewage_siphon,
               "Toilet connection": Sewage_Toilet_connect,
               "Hoses and bolts": Hoses_bolts,
               "Toilets": Toilets,
               "Toilet parts": Toilet_parts_water,
               "Consumables": Consumables,
               "Faucets": Faucets,
               "Sinks": Sinks,
               "Shower sets": Shower_sets,
               "Valves": Valves_water,
               "Rubber seals": Rubber_seals_water,
               "Filters and meters": Filters_meters,
               "Miscellaneous": Miscellaneous_water,
               "Pipe clams": Pipe_clams_water,
               "Water heaters": Water_heaters,
               "Press fittings": Press_fittings,
               "Press pipes": Press_pipes,
               "Screws": Screws,
               "PE pipe": PE_pipe,
               "Radiators": Radiators,
               "Heating valves": Heating_valves,
               "Thermo-manometers": Thermo_manometers,
               "Heat consumables": Heat_consumables,
               "Fiberglass insulation": Fiberglass_insulation,
               "Expansion tanks": Expansion_tanks,
               "Exp tank mounts": Exp_tank_mounts,
               "Manifold parts": Manifold_parts,
               "Heating manifold": Heating_manifold,
               "Gas water heater": Gas_water_heater,
               "Gas heater parts": Gas_heater_parts,
               "Gas fittings": Gas_fittings,
               "Coolant": Coolant,
               "Plastic fittings": Fittings_plastic,
               "Plastic reductions": Reductions_plastic,
               "Modular plastic": Modular_plastic,
               "Pipes": Pipes_plastic,
               "Polyethylene insulation": Polyethylene_isolation,
               "PVC pipes": PVC_pipes,
               "PVC fittings": PVC_fittings,
               "PE fittings": PE_fittings,
               "Bronze fittings": Fittings_bronze,
               "Black metal": Black_metal,
               "Stainless steel": Stainless_steel,
               "Flanges": Flanges,
               "Zinc fittings": Zinc_fittings,
               "Stock plastic fittings": Plastic_stock,
               "Modular plastic fittings": Modular_plastic_stock,
               "Pipe brackets": Brackets_stock,
               "Other": Other_stock}


#===============================================================================================================================================#
# 3. Generator:


def Probability(num):
    probability = np.random.rand(num)
    values = probability / sum(probability)
    return values


def Product_price_generator(
    value,
    price_list,
    higher_limit,
    lower_limit,
    mode='normal',
    custom_weights=None,
    max_parts=None,
    group_map=None,
    Product_names=None,
    min_prob=0.03,
    attempts=30,
    min_quantity=0,
    spread_factor=0.1):
    
    Values = price_list
    n = len(Values)

    for attempt in range(attempts):
        if mode == 'normal':
            Probs = Probability(n)
        elif mode == 'common':
            if custom_weights:
                base_probs = np.array(custom_weights)
            else:
                base_probs = np.maximum(1 / np.array(Values), min_prob)
                base_probs = base_probs / base_probs.sum()

            if group_map and Product_names:
                mask = np.zeros(n)
                for group, limit in group_map.items():
                    group_indices = [i for i, name in enumerate(Product_names) if name == group]
                    if not group_indices:
                        continue
                    group_probs = base_probs[group_indices]
                    group_probs = group_probs / group_probs.sum()
                    selected = np.random.choice(group_indices, size=min(limit, len(group_indices)), replace=False, p=group_probs)
                    for idx in selected:
                        mask[idx] = 1
                Probs = base_probs * mask
                if Probs.sum() == 0:
                    Probs = base_probs
                Probs = Probs / Probs.sum()

            elif max_parts is not None and max_parts < n:
                selected_indices = np.random.choice(range(n), size=max_parts, replace=False, p=base_probs)
                mask = np.zeros(n)
                mask[selected_indices] = 1
                Probs = base_probs * mask
                Probs = Probs / Probs.sum()
            else:
                Probs = base_probs
        else:
            raise ValueError("Mode must be 'normal' or 'common'")

        
        Goal_limits = value * Probs
        Higher_limit = abs(higher_limit / value)
        Lower_limit = abs(lower_limit / value)
        Goal_adjusted = Goal_limits * np.random.uniform(Lower_limit, Higher_limit)
        Quantity = [int(Goal_adjusted[i] // Values[i]) for i in range(n)]
        Final_sum = [Quantity[i] * Values[i] for i in range(n)]
        Results = sum(Final_sum)

        # Responsible for minimum sum target meeting guarantee
        if Results < lower_limit:
            remaining_budget = lower_limit - Results
            available_indices = [i for i in range(n) if Values[i] > 0]
            
            if available_indices:
                weights = [Probs[i] for i in available_indices]
                weights = [w/sum(weights) for w in weights]
                
                items_to_fill = max(1, int(len(available_indices) * spread_factor))
                selected_indices = np.random.choice(
                    available_indices, 
                    size=min(items_to_fill, len(available_indices)), 
                    replace=False, 
                    p=weights)
                
                budget_per_item = remaining_budget / len(selected_indices)
                for idx in selected_indices:
                    additional_quantity = max(min_quantity, int(budget_per_item // Values[idx]))
                    if additional_quantity > 0:
                        Quantity[idx] += additional_quantity
                        remaining_budget -= additional_quantity * Values[idx]
                
                if remaining_budget > 0:
                    while remaining_budget > min(Values) and len(available_indices) > 0:
                        random_idx = np.random.choice(available_indices)
                        if Values[random_idx] <= remaining_budget:
                            additional_qty = max(1, int(remaining_budget // Values[random_idx]))
                            Quantity[random_idx] += additional_qty
                            remaining_budget -= additional_qty * Values[random_idx]
                        else:
                            available_indices.remove(random_idx)
        
        Final_sum = [Quantity[i] * Values[i] for i in range(n)]
        Results = sum(Final_sum)

        if Results >= lower_limit:
            return round(Results, 2), Quantity, Final_sum

    return round(Results, 2), Quantity, Final_sum


def Disassembler(dataframe):
    output = dataframe.to_dict(orient="list")
    output["Index"] = dataframe.index.tolist()
    return output

    
def Invoice_assembler(dataframe, mode='normal', max_parts=None, min_quantity=0, spread_factor=0.1):
    dataframe_copy = dataframe.copy()
    Disassembled_dict = Disassembler(dataframe)
    
    PQ_list = Product_price_generator(
        Price_value,
        Disassembled_dict["Final_price"],
        Lower_limit_value,
        Upper_limit_value,
        mode=mode,
        max_parts=max_parts,
        min_quantity=min_quantity,
        spread_factor=spread_factor)

    dataframe_copy["Quantity"] = PQ_list[1]
    dataframe_copy["Quantity sum"] = PQ_list[2]
    dataframe_copy["Quantity sum"] = dataframe_copy["Quantity sum"].round(2)
    Filtered_data = dataframe_copy[dataframe_copy["Quantity"] > 0]
    return Filtered_data




def simple_Invoice_assembler(dataframe, min_quantity=0, spread_factor=0.1):
    dataframe_copy = dataframe.copy()
    Disassembled_dict = Disassembler(dataframe)
    
    PQ_list = Product_price_generator(
        Price_value, 
        Disassembled_dict["Final_price"],                                  
        Lower_limit_value, 
        Upper_limit_value,
        min_quantity=min_quantity,
        spread_factor=spread_factor)
    
    dataframe_copy["Quantity"] = PQ_list[1]
    dataframe_copy["Quantity sum"] = PQ_list[2]
    dataframe_copy["Quantity sum"] = dataframe_copy["Quantity sum"].round(2)
    Filtered_data = dataframe_copy[dataframe_copy["Quantity"] > 0]
    return Filtered_data



#===============================================================================================================================================#
# 4. Filtering

# for sewage pipes:
Filter_sewage_110 = re.compile(r'(?<!\d)110(?!\d)')
Filter_sewage_50 = re.compile(r'(?<!\d)50(?!\d)')
Filter_sewage_40 = re.compile(r'(?<!\d)40(?!\d)')
Filter_sewage_32 = re.compile(r'(?<!\d)32(?!\d)')


# For diamter:
Filter_20 = re.compile(r'^(?!.*(?:mazgas|Vamzdis)).*(?<!\()20(?!\))(?=\s|x|mm|$)', re.IGNORECASE)
Filter_25 = re.compile(r'^(?!.*(?:mazgas|Vamzdis)).*(?<!\()25(?!\))(?=\s|x|mm|$)', re.IGNORECASE)
Filter_32 = re.compile(r'^(?!.*(?:mazgas|Vamzdis)).*(?<!\()32(?!\))(?=\s|x|mm|$)', re.IGNORECASE)
Filter_40 = re.compile(r'^(?!.*(?:mazgas|Vamzdis)).*(?<!\()40(?!\))(?=\s|x|mm|$)', re.IGNORECASE)
Filter_50 = re.compile(r'^(?!.*(?:mazgas|Vamzdis)).*(?<!\()40(?!\))(?=\s|x|mm|$)', re.IGNORECASE)


# For pipes:
Filter_pipe_hot = re.compile(r'(?=.*(?:Glass|STIK))', re.IGNORECASE)
Filter_pipe_cold = re.compile(r'^(?!.*(?:Glass|STIK))', re.IGNORECASE)

Filter_pipe_20 = re.compile(r'(?<!\d)(?<!PN)20(?!\d)')
Filter_pipe_25 = re.compile(r'(?<!\d)(?<!PN)25(?!\d)')
Filter_pipe_32 = re.compile(r'(?<!\d)(?<!PN)32(?!\d)')
Filter_pipe_40 = re.compile(r'(?<!\d)(?<!PN)40(?!\d)')


# For PVC pipes:
Filter_PVC_20 = re.compile(r'(?<!\d)(?<!PN)(?<!DN)20(?!\d)')
Filter_PVC_25 = re.compile(r'(?<!\d)(?<!PN)(?<!DN)25(?!\d)')
Filter_PVC_32 = re.compile(r'(?<!\d)(?<!PN)(?<!DN)32(?!\d)')
Filter_PVC_50 = re.compile(r'(?<!\d)(?<!PN)(?<!DN)50(?!\d)')


# For Valves:
Valves_1_2 = re.compile(r'(?<!\d\s)(?<!\d")1/2(?!\d)') 
Valves_3_4 = re.compile(r'(?<!\d\s)(?<!\d")3/4(?!\d)')
Valves_1 = re.compile(r'\b1"(?!\d)')
Valves_1_1_4 = re.compile(r'(?<!\d)(?:1 ?"? ?1/4"?(?!\d))')
Valves_1_1_2 = re.compile(r'(?<!\d)(?:1 ?"? ?1/2"?(?!\d))')
Valves_2 = re.compile(r'(?<!\d\/)\b2"(?!\d)')
Valves_2_1_4 = re.compile(r'(?<!\d)(?:2 ?"? ?1/4"?(?!\d))')
Valves_2_1_2 = re.compile(r'(?<!\d)(?:2 ?"? ?1/2"?(?!\d))')


# For sewage pipe clams:
Sewage_clams_1_2 = re.compile(r'(?<![\d"])1/2(?![\d"])|(?:1[5-9]|2[0-7])\s*-\s*(?:2[0-7])|(?:(?<=\s)|^)(?:1[5-9]|2[0-5])(?=\s|$)')
Sewage_clams_3_4 = re.compile(r'(?<![\d"])3/4(?![\d"])|(?:2[5-9])\s*-\s*(?:2[5-9]|3[0-1])|(?:(?<=\s)|^)(?:2[5-9]|3[0-1])(?=\s|$)')
Sewage_clams_1 = re.compile(r'(?<![\d"])1"(?![\d"])|(?:3[0-3])\s*-\s*(?:3[5-7])|(?:(?<=\s)|^)(?:3[0-7])(?=\s|$)')
Sewage_clams_1_1_4 = re.compile(r'(?<![\d"])1 1/4(?![\d])|(?:3[7-9])\s*-\s*(?:4[1-4])|(?:(?<=\s)|^)(?:3[7-9]|4[1-4])(?=\s|$)')
Sewage_clams_1_1_2 = re.compile(r'(?<![\d"])1 1/2(?![\d])|(?:4[6-9])\s*-\s*(?:5[0-3])|(?:(?<=\s)|^)(?:4[6-9]|5[0-3])(?=\s|$)')
Sewage_clams_4 = re.compile(r'(?<![\d"|/])4"(?![\d])|(?:11[1-5])\s*-\s*(?:12[0-5])|(?:(?<=\s)|^)(?:11\d|12[0-5])(?=\s|$)')


# For Chrome:
Chrome_1_2 = re.compile(r'(?<!\d\s)(?<!\d")1/2(?!\d)') 
Chrome_3_4 = re.compile(r'(?<!\d\s)(?<!\d")3/4(?!\d)')


# For Toilets:
Toilet_wc = re.compile(r'(W00(0[0-9]|[1-9][0-9]))$')


# For Water heaters:
Water_heater_option = re.compile(r'WH0(?:[0-9][0-9])$')


# For press fittings:
Filter_press_multilayer = re.compile(r'^(?!.*(?:Press|therm|pl\.))(?=.*(?:16|18|20|26|32))', re.IGNORECASE)
Filter_press_metal = re.compile(r'^(?=.*(?:18|22|28|35|42))(?=.*(?:cink|KAN|Plien))', re.IGNORECASE)


# For press diameter:
Filter_press_16 = re.compile(r'(?<!\d)16(?!\d)')
Filter_press_18 = re.compile(r'(?<!\d)18(?!\d)')
Filter_press_20 = re.compile(r'(?<!\d)20(?!\d)')
Filter_press_22 = re.compile(r'(?<!\d)22(?!\d)')
Filter_press_26 = re.compile(r'(?<!\d)26(?!\d)')
Filter_press_28 = re.compile(r'(?<!\d)28(?!\d)')
Filter_press_32 = re.compile(r'(?<!\d)32(?!\d)')
Filter_press_35 = re.compile(r'(?<!\d)35(?!\d)')
Filter_press_42 = re.compile(r'(?<!\d)42(?!\d)')


# For Radiators:
Radiators_only = re.compile(r'^R00(?:[0-9][0-9])')
Radiator_parts = re.compile(r'^R0(?:[1-2][0-9][0-9])$')


# For Radiators:
Valves_only = re.compile(r'(?=.*(?:ventilis))', re.IGNORECASE)
Air_valves = re.compile(r'(?=.*(?:nuorin))', re.IGNORECASE)




#===============================================================================================================================================#
# 5. Interface:

Output = widgets.Output()
Filter_output = widgets.Output()
Filter_inches_output = widgets.Output()
Filter_pipe_output = widgets.Output()
Filter_pipe_diameter_output = widgets.Output()
Filter_pipe_sewage_diameter_output = widgets.Output()
Filter_sewage_pipe_clams_output = widgets.Output()
Filter_chrome_output = widgets.Output()
Filter_toilet_output = widgets.Output()
Filter_heater_output = widgets.Output()
Filter_press_pipe_output = widgets.Output()
Filter_press_diameter_output = widgets.Output()
Filter_PVC_pipe_output = widgets.Output()
Filter_radiators_output = widgets.Output()
Filter_heat_valves_output = widgets.Output()
Combined_output = widgets.Output()

# Data selection:
Category_checkboxes = {
    "Chrome fittings": widgets.Checkbox(value=False, description="Chrome fittings", layout=widgets.Layout(margin="0 0 0 -80px", width="250px")),
    "Sewage fittings": widgets.Checkbox(value=False, description="Sewage fittings", layout=widgets.Layout(margin="0 0 0 -80px", width="250px")),
    "Water fittings": widgets.Checkbox(value=False, description="Water fittings", layout=widgets.Layout(margin="0 0 0 -80px", width="250px")),
    "Heat fittings": widgets.Checkbox(value=False, description="Heating fittings", layout=widgets.Layout(margin="0 0 0 -80px", width="250px")),
    "Plastic fittings": widgets.Checkbox(value=False, description="Plastic fittings", layout=widgets.Layout(margin="0 0 0 -80px",width="250px")),
    "Bronze fittings": widgets.Checkbox(value=False, description="Bronze fittings", layout=widgets.Layout(margin="0 0 0 -80px",width="250px")),
    "Metal fittings": widgets.Checkbox(value=False, description="Metal fittings", layout=widgets.Layout(margin="0 0 0 -80px",width="250px")),
    "Stock": widgets.Checkbox(value=False, description="Stock", layout=widgets.Layout(margin="0 0 0 -80px",width="250px"))}

Category_box = widgets.VBox(list(Category_checkboxes.values()))

Checkbox_dictionary = {"Chrome fittings": ["Chrome parts", "Extenders"],
                       "Sewage fittings": ["Sewage pipes", "Pumps", "Sewage clams", "Chemicals", "Siphons", "Toilet connection"],
                       "Water fittings": ["Hoses and bolts", "Toilets", "Toilet parts", "Consumables", "Faucets", "Sinks", "Shower sets",\
                                          "Valves", "Rubber seals", "Filters and meters", "Miscellaneous", "Pipe clams", "Water heaters",\
                                          "Press fittings", "Press pipes", "Screws", "PE pipe"],
                       "Heat fittings": ["Radiators", "Heating valves", "Thermo-manometers", "Heat consumables", "Fiberglass insulation",\
                                         "Expansion tanks", "Exp tank mounts", "Manifold parts", "Heating manifold", "Gas water heater",\
                                         "Gas heater parts", "Gas fittings", "Coolant"],
                       "Plastic fittings":["Plastic fittings", "Plastic reductions", "Modular plastic", "Pipes", "Polyethylene insulation",\
                                           "PVC pipes", "PVC fittings", "PE fittings"],
                       "Bronze fittings": ["Bronze fittings"],
                       "Metal fittings": ["Black metal", "Stainless steel", "Flanges", "Zinc fittings"],
                       "Stock": ["Stock plastic fittings", "Modular plastic fittings", "Pipe brackets", "Other"]}

Item_checkboxes = {}
Item_box = widgets.VBox()


def Update_items(*args):
    global Item_checkboxes
    Item_checkboxes = {}
    checkers = []

    for category, checkbox in Category_checkboxes.items():
        if checkbox.value:
            for item in Checkbox_dictionary[category]:
                cb = widgets.Checkbox(value=False, description=item, layout = widgets.Layout(margin="0 0 0 -85px", width="200px"))
                Item_checkboxes[item] = cb
                checkers.append(cb)

    max_per_column = 7
    columns = [widgets.VBox(checkers[i:i + max_per_column])
                for i in range(0, len(checkers), max_per_column)]
    Item_box.children = [widgets.HBox(columns)]

    for cb in Item_checkboxes.values():
        cb.observe(Filter_updater, "value")
        cb.observe(Filter_pipe_updater, "value")
        cb.observe(Filter_inches_updater, "value")
        cb.observe(Filter_sewage_diameter_updater, "value")
        cb.observe(Filter_sewage_clam_updater, "value")
        cb.observe(Filter_chrome_updater, "value")
        cb.observe(Filter_toilet_updater, "value")
        cb.observe(Filter_heater_updater, "value")
        cb.observe(Filter_press_pipe_updater, "value")
        cb.observe(Filter_PVC_pipe_updater, "value")
        cb.observe(Filter_radiator_updater, "value")
        cb.observe(Filter_heat_valve_updater, "value")
        
    update_group_common_filters()


for cb in Category_checkboxes.values():
    cb.observe(Update_items, "value")


#===============================================================================================================================================#
# 6. Filter selection:

Filter_boxes = {
    "D20": widgets.Checkbox(value=False, description = "D20 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="200px")),
    "D25": widgets.Checkbox(value=False, description = "D25 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="200px")),
    "D32": widgets.Checkbox(value=False, description = "D32 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="200px")),
    "D40": widgets.Checkbox(value=False, description = "D40 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="200px"))}
    
Filter_inches_boxes = {
    "1/2": widgets.Checkbox(value=False, description = "1/2 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="220px")),
    "3/4": widgets.Checkbox(value=False, description = "3/4 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="220px")),
    '1"': widgets.Checkbox(value=False, description = '1" Diameter', layout = widgets.Layout(margin="0 0 0 -85px", width="220px")),
    '1 1/4"': widgets.Checkbox(value=False, description = '1 1/4" Diameter', layout = widgets.Layout(margin="0 0 0 -85px", width="220px")),
    '1 1/2"': widgets.Checkbox(value=False, description = '1 1/2" Diameter', layout = widgets.Layout(margin="0 0 0 -85px", width="220px")),
    '2"': widgets.Checkbox(value=False, description = '2" Diameter', layout = widgets.Layout(margin="0 0 0 -85px", width="220px")),
    '2 1/4"': widgets.Checkbox(value=False, description = '2 1/4" Diameter', layout = widgets.Layout(margin="0 0 0 -85px", width="220px")),
    '2 1/2"': widgets.Checkbox(value=False, description = '2 1/2" Diameter', layout = widgets.Layout(margin="0 0 0 -85px", width="220px"))}

Filter_pipe_boxes = {
    "Hot": widgets.Checkbox(value=False, description = "Hot water", layout = widgets.Layout(margin="0 0 0 -85px", width="200px")),
    "Cold": widgets.Checkbox(value=False, description = "Cold water", layout = widgets.Layout(margin="0 0 0 -85px", width="200px"))}

Filter_pipe_diameter_boxes = {
    "D20 pipe": widgets.Checkbox(value=False, description = "D20 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="200px")),
    "D25 pipe": widgets.Checkbox(value=False, description = "D25 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="200px")),
    "D32 pipe": widgets.Checkbox(value=False, description = "D32 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="200px")),
    "D40 pipe": widgets.Checkbox(value=False, description = "D40 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="200px"))}

Filter_sewage_pipe_diamter_boxes = {
   "D32_sewage": widgets.Checkbox(value=False, description = "D32 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="200px")),
   "D40_sewage": widgets.Checkbox(value=False, description = "D40 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="200px")),
   "D50_sewage": widgets.Checkbox(value=False, description = "D50 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="200px")),
   "D110_sewage": widgets.Checkbox(value=False, description = "D110 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="200px"))}

Filter_sewage_pipe_clams_boxes = {
    "1/2_sew": widgets.Checkbox(value=False, description = "1/2 (20 mm)", layout = widgets.Layout(margin="0 0 0 -85px", width="220px")),
    "3/4_sew": widgets.Checkbox(value=False, description = "3/4 (25 mm)", layout = widgets.Layout(margin="0 0 0 -85px", width="220px")),
    '1"_sew': widgets.Checkbox(value=False, description = '1" (32 mm)', layout = widgets.Layout(margin="0 0 0 -85px", width="220px")),
    '1 1/4"_sew': widgets.Checkbox(value=False, description = '1 1/4" (40 mm)', layout = widgets.Layout(margin="0 0 0 -85px", width="220px")),
    '1 1/2"_sew': widgets.Checkbox(value=False, description = '1 1/2" (50 mm)', layout = widgets.Layout(margin="0 0 0 -85px", width="220px")),
    '4"_sew': widgets.Checkbox(value=False, description = '4" (110 mm)', layout = widgets.Layout(margin='0 0 0 -85px', width="220px"))}

Filter_chrome_boxes = {
    "1/2_chrom": widgets.Checkbox(value=False, description = "1/2 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="220px")),
    "3/4_chrom": widgets.Checkbox(value=False, description = "3/4 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="220px"))}

Filter_toilet_boxes = {
    "Only toilets": widgets.Checkbox(value=False, description = "Only toilets", layout = widgets.Layout(margin="0 0 0 -85px", width="220px"))}

Filter_heater_boxes = {
    "Only heaters": widgets.Checkbox(value=False, description = "Only heaters", layout = widgets.Layout(margin="0 0 0 -85px", width="220px"))}

Filter_press_pipe_boxes = {
    "Multilayer": widgets.Checkbox(value=False, description = "Multilayer pipe", layout = widgets.Layout(margin="0 0 0 -85px", width="220px")),
    "Metal": widgets.Checkbox(value=False, description = "Metal pipe", layout = widgets.Layout(margin="0 0 0 -85px", width="220px"))}

Filter_press_diameter_boxes = {
    "D16_Press": widgets.Checkbox(value=False, description = "16 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="220px")),
    "D18_Press": widgets.Checkbox(value=False, description = "18 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="220px")),
    "D20_Press": widgets.Checkbox(value=False, description = "20 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="220px")),
    "D22_Press": widgets.Checkbox(value=False, description = "22 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="220px")),
    "D26_Press": widgets.Checkbox(value=False, description = "26 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="220px")),
    "D28_Press": widgets.Checkbox(value=False, description = "28 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="220px")),
    "D32_Press": widgets.Checkbox(value=False, description = "32 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="220px")),
    "D35_Press": widgets.Checkbox(value=False, description = "35 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="220px")),
    "D42_Press": widgets.Checkbox(value=False, description = "42 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="220px"))}

Filter_PVC_pipe_boxes = {
    "D20_pvc": widgets.Checkbox(value=False, description = "20 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="220px")),
    "D25_pvc": widgets.Checkbox(value=False, description = "25 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="220px")),
    "D32_pvc": widgets.Checkbox(value=False, description = "32 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="220px")),
    "D50_pvc": widgets.Checkbox(value=False, description = "50 Diameter", layout = widgets.Layout(margin="0 0 0 -85px", width="220px"))}

Filter_radiator_boxes = {
    "Only radiator": widgets.Checkbox(value=False, description = "Only radiators", layout = widgets.Layout(margin="0 0 0 -85px", width="220px")),
    "Only rad parts": widgets.Checkbox(value=False, description = "Only parts", layout = widgets.Layout(margin="0 0 0 -85px", width="220px"))}

Filter_heat_valves_boxes = {
    "Heat valves": widgets.Checkbox(value=False, description = "Only valves", layout = widgets.Layout(margin="0 0 0 -85px", width="220px")),
    "Air valves": widgets.Checkbox(value=False, description = "Only air valves", layout = widgets.Layout(margin="0 0 0 -85px", width="220px"))}


Filter_match = {"D20": Filter_20,
                "D25": Filter_25,
                "D32": Filter_32,
                "D40": Filter_40,
                "Hot": Filter_pipe_hot,
                "Cold": Filter_pipe_cold,
                "D20 pipe": Filter_pipe_20,
                "D25 pipe": Filter_pipe_25,
                "D32 pipe": Filter_pipe_32,
                "D40 pipe": Filter_pipe_40,
                "1/2": Valves_1_2,
                "3/4": Valves_3_4,
                '1"': Valves_1,
                '1 1/4"': Valves_1_1_4,
                '1 1/2"': Valves_1_1_2,
                '2"': Valves_2,
                '2 1/4"': Valves_2_1_4,
                '2 1/2"': Valves_2_1_2,
                "D32_sewage": Filter_sewage_32,
                "D40_sewage": Filter_sewage_40,
                "D50_sewage": Filter_sewage_50,
                "D110_sewage": Filter_sewage_110,
                "1/2_sew": Sewage_clams_1_2,
                "3/4_sew": Sewage_clams_3_4,
                '1"_sew': Sewage_clams_1,
                '1 1/4"_sew': Sewage_clams_1_1_4,
                '1 1/2"_sew': Sewage_clams_1_1_2,
                '4"_sew': Sewage_clams_4,
                "1/2_chrom": Chrome_1_2,
                "3/4_chrom": Chrome_3_4,
                "Only toilets": Toilet_wc,
                "Only heaters": Water_heater_option,
                "Multilayer": Filter_press_multilayer,
                "Metal": Filter_press_metal,
                "D16_Press": Filter_press_16,
                "D18_Press": Filter_press_18,
                "D20_Press": Filter_press_20,
                "D22_Press": Filter_press_22,
                "D26_Press": Filter_press_26,
                "D28_Press": Filter_press_28,
                "D32_Press": Filter_press_32,
                "D35_Press": Filter_press_35,
                "D42_Press": Filter_press_42,
                "D20_pvc": Filter_PVC_20,
                "D25_pvc": Filter_PVC_25,
                "D32_pvc": Filter_PVC_32,
                "D50_pvc": Filter_PVC_50,
                "Only radiator": Radiators_only,
                "Only rad parts": Radiator_parts,
                "Heat valves": Valves_only,
                "Air valves": Air_valves}

Filter_box = widgets.VBox(list(Filter_boxes.values()))
Filter_inches_box = widgets.VBox(list(Filter_inches_boxes.values()))
Filter_pipe_box = widgets.VBox(list(Filter_pipe_boxes.values()))
Filter_pipe_diameter_box = widgets.VBox(list(Filter_pipe_diameter_boxes.values()))
Filter_sewage_pipe_diameter_box = widgets.VBox(list(Filter_sewage_pipe_diamter_boxes.values()))
Filter_sewage_clam_diameter_box = widgets.VBox(list(Filter_sewage_pipe_clams_boxes.values()))
Filter_chrome_box = widgets.VBox(list(Filter_chrome_boxes.values()))
Filter_toilet_box = widgets.VBox(list(Filter_toilet_boxes.values()))
Filter_heater_box = widgets.VBox(list(Filter_heater_boxes.values()))
Filter_press_pipe_box = widgets.VBox(list(Filter_press_pipe_boxes.values()))
Filter_press_diameter_box = widgets.VBox(list(Filter_press_diameter_boxes.values()))
Filter_PVC_pipe_box = widgets.VBox(list(Filter_PVC_pipe_boxes.values()))
Filter_radiator_box = widgets.VBox(list(Filter_radiator_boxes.values()))
Filter_heat_valves_box = widgets.VBox(list(Filter_heat_valves_boxes.values()))


###############################################################
###################### Filter assigning #######################
###############################################################

Checkbox_dictionary_Filter = {"Chrome parts": ["1/2_chrom", "3/4_chrom"],
                              "Extenders": ["1/2_chrom"],
                              "Sewage pipes": ["D32_sewage", "D40_sewage", "D50_sewage", "D110_sewage"],
                              "Pumps": None,
                              "Sewage clams": ["1/2_sew", "3/4_sew", '1"_sew', '1 1/4"_sew', '1 1/2"_sew', '4"_sew'],
                              "Chemicals": None,
                              "Siphons": None,
                              "Toilet connection": None,
                              "Hoses and bolts": None,
                              "Toilets": ["Only toilets"],
                              "Toilet parts": None,
                              "Consumables": None,
                              "Faucets": None,
                              "Sinks": None,
                              "Shower sets": None,
                              "Valves": ["1/2", "3/4", '1"', '1 1/4"', '1 1/2"', '2"', '2 1/2"'],
                              "Rubber seals": ["1/2", "3/4", '1"', '1 1/4"', '1 1/2"', '2"', '2 1/2"'],
                              "Filters and meters": None,
                              "Miscellaneous": None,
                              "Pipe clams": "D25",
                              "Water heaters": ["Only heaters"],
                              "Press fittings": ["Multilayer", "Metal"],
                              "Press pipes": ["Multilayer", "Metal"],
                              "Screws": None,
                              "PE pipe": None,
                              "Radiators": ["Only radiator", "Only rad parts"],
                              "Heating valves": ["Heat valves", "Air valves"],
                              "Thermo-manometers": None,
                              "Heat consumables": None,
                              "Fiberglass insulation": None,
                              "Expansion tanks": None,
                              "Exp tank mounts": None,
                              "Manifold parts": None,
                              "Heating manifold": None,
                              "Gas water heater": None,
                              "Gas heater parts": None,
                              "Gas fittings": None, 
                              "Coolant": None,
                              "Plastic fittings": ["D20", "D25", "D32", "D40"],
                              "Plastic reductions": ["D20", "D25", "D32", "D40"],
                              "Modular plastic": ["D20", "D25", "D32", "D40"],
                              "Pipes": ["Hot", "Cold"],
                              "Polyethylene insulation": None,
                              "PVC pipes": ["D20_pvc", "D25_pvc", "D50_pvc"],
                              "PVC fittings": ["D20_pvc", "D25_pvc", "D32_pvc", "D50_pvc"],
                              "PE fittings": ["D20", "D25", "D32"],
                              "Bronze fittings": ["1/2", "3/4", '1"', '1 1/4"', '1 1/2"'],
                              "Black metal": ["1/2", "3/4", '1"', '1 1/4"', '1 1/2"', '2"', '2 1/2"'],
                              "Stainless steel": None,
                              "Flanges": None,
                              "Zinc fittings": ["3/4", '1"', '1 1/4"', '1 1/2"'],
                              "Stock plastic fittings": ["D20", "D25", "D32", "D40"],
                              "Modular plastic fittings": ["D20", "D25", "D32", "D40"],
                              "Pipe brackets": ["D25", "D32"],
                              "Other": None}

Pipe_dictionary_Filter = {"Hot": ["D20 pipe", "D25 pipe", "D32 pipe", "D40 pipe"],
                          "Cold": ["D20 pipe", "D25 pipe", "D32 pipe", "D40 pipe"]}

Press_pipe_dictionary_Filter = {"Multilayer": ["D16_Press", "D18_Press", "D20_Press", "D26_Press", "D32_Press"],
                                "Metal": ["D22_Press", "D28_Press", "D35_Press", "D42_Press"]}


#===============================================================================================================================================#
# 7. Common items and sliders:

common_items_toggle = widgets.Checkbox(value=False, description="Use common items", layout=widgets.Layout(width="250px", margin="0 0 0 -70px"))
max_parts_slider = widgets.IntSlider(value=3, min=1, max=10, step=1, description='Max parts:', layout=widgets.Layout(width='300px'))


def update_max_parts_slider(change=None):
    """
    Update the max_parts_slider's max value for each group based on current filters
    """
    with Output:
        if not common_items_toggle.value:
            if hasattr(Invoice_filtering, 'last_filtered_df'): # <-- reference point is created in Invoice_filtering (True/False)
                filtered_count = len(Invoice_filtering.last_filtered_df)
                max_parts_slider.max = max(1, filtered_count)
                if max_parts_slider.value > max_parts_slider.max:
                    max_parts_slider.value = max_parts_slider.max
        else:
            for widget_box in group_common_filter_box.children:
                checkbox, slider_box = widget_box.children
                group_name = checkbox.description
                slider = slider_box.children[1]
                
                if group_name in Combined_df:
                    df = Combined_df[group_name].copy()
                    mask = pd.Series([True] * len(df), index=df.index)

                    if group_name in ["Valves", "Rubber seals", "Bronze fittings", "Black metal", "Zinc fittings"]:
                        inches_mask = pd.Series([False] * len(df), index=df.index)
                        for inch_key, cb in Filter_inches_boxes.items():
                            if cb.value:
                                regex_inches = Filter_match.get(inch_key, "")
                                inches_mask |= df["Name"].str.contains(regex_inches)
                        if inches_mask.any():
                            mask &= inches_mask
                  
                    elif group_name in ["Sewage pipes"]:
                        sewage_mask = pd.Series([False] * len(df), index=df.index)
                        for sewage_diameter, cb in Filter_sewage_pipe_diamter_boxes.items():
                            if cb.value:
                                regex_sewage = Filter_match.get(sewage_diameter, "")
                                sewage_mask |= df["Name"].str.contains(regex_sewage)
                        if sewage_mask.any():
                            mask &= sewage_mask
                    
                    elif group_name in ["Sewage clams"]:
                        sewage_clam_mask = pd.Series([False] * len(df), index=df.index)
                        for clam_diameter, cb in Filter_sewage_pipe_clams_boxes.items():
                            if cb.value:
                                regex_sewage_clam = Filter_match.get(clam_diameter, "")
                                sewage_clam_mask |= df["Name"].str.contains(regex_sewage_clam)
                        if sewage_clam_mask.any():
                            mask &= sewage_clam_mask

                    elif group_name in ["Chrome parts", "Extenders"]:
                        chrome_mask = pd.Series([False] * len(df), index=df.index)
                        for chrome_diameter, cb in Filter_chrome_boxes.items():
                            if cb.value:
                                regex_chrome = Filter_match.get(chrome_diameter, "")
                                chrome_mask |= df["Name"].str.contains(regex_chrome)
                        if chrome_mask.any():
                            mask &= chrome_mask

                        
                    elif group_name in ["Toilets"]:
                        toilet_mask = pd.Series([False] * len(df), index=df.index)
                        for toilet_check, cb in Filter_toilet_boxes.items():
                            if cb.value:
                                regex_toilet = Filter_match.get(toilet_check, "")
                                toilet_mask |= df.index.str.contains(regex_toilet)
                        if toilet_mask.any():
                            mask &= toilet_mask
                              
                        
                    elif group_name in ["Water heaters"]:
                        heater_mask = pd.Series([False] * len(df), index=df.index)
                        for heater_check, cb in Filter_heater_boxes.items():
                            if cb.value:
                                regex_heater = Filter_match.get(heater_check, "")
                                heater_mask |= df.index.str.contains(regex_heater)
                        if heater_mask.any():
                            mask &= heater_mask
                            
                            
                    elif group_name in ["Pipes", "Plastic fittings", "Plastic reductions", "Modular plastic",\
                                        "Stock plastic fittings", "Modular plastic fittings", "Pipe brackets",\
                                        "PE fittings"]:
                        diam_mask = pd.Series([False] * len(df), index=df.index)
                        if group_name == "Pipes":
                            for diam_key, cb in Filter_pipe_diameter_boxes.items():
                                if cb.value:
                                    regex_pipe = Filter_match.get(diam_key, "")
                                    diam_mask |= df["Name"].str.contains(regex_pipe)
                        else:
                            for fitting_key, cb in Filter_boxes.items():
                                if cb.value:
                                    regex_diam = Filter_match.get(fitting_key, "")
                                    diam_mask |= df["Name"].str.contains(regex_diam)
                        if diam_mask.any():
                            mask &= diam_mask

                    
                    elif group_name in ["Press fittings", "Press pipes"]: 
                        Press_mask = pd.Series([False] * len(df), index=df.index)
                        for press_check, cb in Filter_press_diameter_boxes.items():
                            if cb.value:
                                regex_press = Filter_match.get(press_check, "")
                                Press_mask |= df["Name"].str.contains(regex_press)
                        if Press_mask.any():
                            mask &= Press_mask

                            
                    elif group_name in ["PVC pipes", "PVC fittings"]: 
                        PVC_mask = pd.Series([False] * len(df), index=df.index)
                        for PVC_check, cb in Filter_PVC_pipe_boxes.items():
                            if cb.value:
                                regex_PVC = Filter_match.get(PVC_check, "")
                                PVC_mask |= df["Name"].str.contains(regex_PVC)
                        if PVC_mask.any():
                            mask &= PVC_mask
                            
                            
                    elif group_name in ["Radiators"]: 
                        Radiator_mask = pd.Series([False] * len(df), index=df.index)
                        for radiator_check, cb in Filter_radiator_boxes.items():
                            if cb.value:
                                regex_radiator = Filter_match.get(radiator_check, "")
                                Radiator_mask |= df.index.str.contains(regex_radiator)
                        if Radiator_mask.any():
                            mask &= Radiator_mask
                            
                            
                    elif group_name in ["Heating valves"]: 
                        heat_valve_mask = pd.Series([False] * len(df), index=df.index)
                        for heat_valve_check, cb in Filter_heat_valves_boxes.items():
                            if cb.value:
                                regex_heat_valve = Filter_match.get(heat_valve_check, "")
                                heat_valve_mask |= df["Name"].str.contains(regex_heat_valve)
                        if heat_valve_mask.any():
                            mask &= heat_valve_mask
                    
                    filtered_count = len(df[mask])
                    slider.max = max(1, filtered_count)
                    if slider.value > slider.max:
                        slider.value = slider.max

max_parts_slider.max = 10

#===============================================================================================================================================#
# 8. Filter updaters:

def Filter_updater(change):
    with Filter_output:
        Filter_output.clear_output()
        Selected_items = [key for key, cb in Item_checkboxes.items() if cb.value]
        active_filter_keys = set()

        for item in Selected_items:
            filter_entry = Checkbox_dictionary_Filter.get(item)
            if isinstance(filter_entry, list):
                active_filter_keys.update(filter_entry)
            elif isinstance(filter_entry, str):
                active_filter_keys.add(filter_entry)

        
        Diameter_filter_order = ["D20", "D25", "D32", "D40"]
        ordered_fitting_diameters = [Filter_boxes[key]for key in Diameter_filter_order if key in active_filter_keys]
        Filter_box.children = ordered_fitting_diameters
        update_top_filter_row()
        update_max_parts_slider()


def Filter_pipe_updater(change):
    Selected_items = [key for key, cb in Item_checkboxes.items() if cb.value]
    active_filter_keys = set()

    for item in Selected_items:
        entry = Checkbox_dictionary_Filter.get(item)
        if isinstance(entry, list):
            active_filter_keys.update(entry)
        elif isinstance(entry, str):
            active_filter_keys.add(entry)

    matching_filters = [Filter_pipe_boxes[key] for key in active_filter_keys if key in Filter_pipe_boxes]
    Filter_pipe_box.children = matching_filters

    for key in active_filter_keys:
        if key in Filter_pipe_boxes:
            Filter_pipe_boxes[key].observe(Filter_pipe_diameter_updater, "value")
        
    Filter_pipe_diameter_updater(None)
    update_max_parts_slider()
   
    
def Filter_pipe_diameter_updater(change):
    selected_pipe_types = [key for key, cb in Filter_pipe_boxes.items() if cb.value]
    active_diameters = set()

    for pipe_type in selected_pipe_types:
        diameters = Pipe_dictionary_Filter.get(pipe_type)
        if isinstance(diameters, list):
            active_diameters.update(diameters)
        elif isinstance(diameters, str):
            active_diameters.add(diameters)


    pipe_diameter_order = ["D20 pipe", "D25 pipe", "D32 pipe", "D40 pipe"]
    ordered_diameters = [Filter_pipe_diameter_boxes[key]for key in pipe_diameter_order if key in active_diameters]

    Filter_pipe_diameter_box.children = ordered_diameters
    update_top_filter_row()
    update_max_parts_slider()


def Filter_inches_updater(change):
    with Filter_inches_output:
        Filter_inches_output.clear_output()
        Selected_items_4 = [key for key, cb in Item_checkboxes.items() if cb.value]
        active_filter_key_4 = set()

        for item_4 in Selected_items_4:
            Filter_inches_entry = Checkbox_dictionary_Filter.get(item_4)
            if isinstance(Filter_inches_entry, list):
                active_filter_key_4.update(Filter_inches_entry)
            elif isinstance(Filter_inches_entry, str):
                active_filter_key_4.add(Filter_inches_entry)

        ordered_filters = []
        for key, cb in Filter_inches_boxes.items():
            if key in active_filter_key_4:
                ordered_filters.append(Filter_inches_boxes[key])
            
        Filter_inches_box.children = ordered_filters
        update_top_filter_row()
        update_max_parts_slider()


def Filter_sewage_diameter_updater(change):
    with Filter_pipe_sewage_diameter_output:
        Filter_pipe_sewage_diameter_output.clear_output()
        Selected_items_5 = [key for key, cb in Item_checkboxes.items() if cb.value]
        active_filter_key_5 = set()

        for item_5 in Selected_items_5:
            sewage_entry = Checkbox_dictionary_Filter.get(item_5)
            if isinstance(sewage_entry, list):
                active_filter_key_5.update(sewage_entry)
            elif isinstance(sewage_entry, str):
                active_filter_key_5.add(sewage_entry)

        sewage_pipe_order = []
        for key, cb in Filter_sewage_pipe_diamter_boxes.items():
            if key in active_filter_key_5:
                sewage_pipe_order.append(cb)

        Filter_sewage_pipe_diameter_box.children = sewage_pipe_order
        update_top_filter_row()
        update_max_parts_slider()
    

def Filter_sewage_clam_updater(change):
    with Filter_sewage_pipe_clams_output:
        Filter_sewage_pipe_clams_output.clear_output()
        Selected_items_6 = [key for key, cb in Item_checkboxes.items() if cb.value]
        active_filter_key_6 = set()

        for item_6 in Selected_items_6:
            sewage_clam_entry = Checkbox_dictionary_Filter.get(item_6)
            if isinstance(sewage_clam_entry, list):
                active_filter_key_6.update(sewage_clam_entry)
            elif isinstance(sewage_clam_entry, str):
                active_filter_key_6.add(sewage_clam_entry)

        sewage_clam_order = []
        for key, cb in Filter_sewage_pipe_clams_boxes.items():
            if key in active_filter_key_6:
                sewage_clam_order.append(cb)

        Filter_sewage_clam_diameter_box.children = sewage_clam_order
        update_top_filter_row()
        update_max_parts_slider()
    

def Filter_chrome_updater(change):
    with Filter_chrome_output:
        Filter_chrome_output.clear_output()
        Selected_items_7 = [key for key, cb in Item_checkboxes.items() if cb.value]
        active_filter_key_7 = set()

        for item_7 in Selected_items_7:
            chrome_entry = Checkbox_dictionary_Filter.get(item_7)
            if isinstance(chrome_entry, list):
                active_filter_key_7.update(chrome_entry)
            elif isinstance(chrome_entry, str):
                active_filter_key_7.add(chrome_entry)

        chrome_order = []
        for key, cb in Filter_chrome_boxes.items():
            if key in active_filter_key_7:
                chrome_order.append(cb)

        Filter_chrome_box.children = chrome_order
        update_top_filter_row()
        update_max_parts_slider()


def Filter_toilet_updater(change):
    with Filter_toilet_output:
        Filter_toilet_output.clear_output()
        Selected_items_8 = [key for key, cb in Item_checkboxes.items() if cb.value]
        active_filter_key_8 = set()

        for item_8 in Selected_items_8:
            toilet_entry = Checkbox_dictionary_Filter.get(item_8)
            if isinstance(toilet_entry, list):
                active_filter_key_8.update(toilet_entry)
            elif isinstance(toilet_entry, str):
                active_filter_key_8.add(toilet_entry)

        toilet_order = []
        for key, cb in Filter_toilet_boxes.items():
            if key in active_filter_key_8:
                toilet_order.append(cb)

        Filter_toilet_box.children = toilet_order
        update_top_filter_row()
        update_max_parts_slider()


def Filter_heater_updater(change):
    with Filter_heater_output:
        Filter_heater_output.clear_output()
        Selected_items_9 = [key for key, cb in Item_checkboxes.items() if cb.value]
        active_filter_key_9 = set()

        for item_9 in Selected_items_9:
            heater_entry = Checkbox_dictionary_Filter.get(item_9)
            if isinstance(heater_entry, list):
                active_filter_key_9.update(heater_entry)
            elif isinstance(heater_entry, str):
                active_filter_key_9.add(heater_entry)

        heater_order = []
        for key, cb in Filter_heater_boxes.items():
            if key in active_filter_key_9:
                heater_order.append(cb)

        Filter_heater_box.children = heater_order
        update_top_filter_row()
        update_max_parts_slider()


def Filter_press_pipe_updater(change):
    Selected_items_10 = [key for key, cb in Item_checkboxes.items() if cb.value]
    active_filter_keys_10 = set()

    for item_10 in Selected_items_10:
        press_entry = Checkbox_dictionary_Filter.get(item_10)
        if isinstance(press_entry, list):
            active_filter_keys_10.update(press_entry)
        elif isinstance(press_entry, str):
            active_filter_keys_10.add(press_entry)

    matching_filters = [Filter_press_pipe_boxes[key] for key in active_filter_keys_10 if key in Filter_press_pipe_boxes]
    Filter_press_pipe_box.children = matching_filters

    for key in active_filter_keys_10:
        if key in Filter_press_pipe_boxes:
            Filter_press_pipe_boxes[key].observe(Filter_press_diameter_updater, "value")
        
    Filter_press_diameter_updater(None)
    update_max_parts_slider()


def Filter_press_diameter_updater(change):
    selected_press_types = [key for key, cb in Filter_press_pipe_boxes.items() if cb.value]
    active_filter_keys_11 = set()

    for press_type in selected_press_types:
        press_diameters = Press_pipe_dictionary_Filter.get(press_type)
        if isinstance(press_diameters, list):
            active_filter_keys_11.update(press_diameters)
        elif isinstance(press_diameters, str):
            active_filter_keys_11.add(press_diameters)


    Press_diameter_order = ["D16_Press", "D18_Press", "D20_Press", "D22_Press", "D26_Press", "D28_Press", "D32_Press", "D35_Press", "D42_Press"]
    ordered_press_diameters = [Filter_press_diameter_boxes[key]for key in Press_diameter_order if key in active_filter_keys_11]

    Filter_press_diameter_box.children = ordered_press_diameters
    update_top_filter_row()
    update_max_parts_slider()


def Filter_PVC_pipe_updater(change):
    with Filter_PVC_pipe_output:
        Filter_PVC_pipe_output.clear_output()
        Selected_items_12 = [key for key, cb in Item_checkboxes.items() if cb.value]
        active_filter_key_12 = set()

        for item_12 in Selected_items_12:
            PVC_pipe_entry = Checkbox_dictionary_Filter.get(item_12)
            if isinstance(PVC_pipe_entry, list):
                active_filter_key_12.update(PVC_pipe_entry)
            elif isinstance(PVC_pipe_entry, str):
                active_filter_key_12.add(PVC_pipe_entry)

        PVC_pipe_order = []
        for key, cb in Filter_PVC_pipe_boxes.items():
            if key in active_filter_key_12:
                PVC_pipe_order.append(cb)

        Filter_PVC_pipe_box.children = PVC_pipe_order
        update_top_filter_row()
        update_max_parts_slider()


def Filter_radiator_updater(change):
    with Filter_radiators_output:
        Filter_radiators_output.clear_output()
        Selected_items_13 = [key for key, cb in Item_checkboxes.items() if cb.value]
        active_filter_key_13 = set()

        for item_13 in Selected_items_13:
            radiator_entry = Checkbox_dictionary_Filter.get(item_13)
            if isinstance(radiator_entry, list):
                active_filter_key_13.update(radiator_entry)
            elif isinstance(radiator_entry, str):
                active_filter_key_13.add(radiator_entry)

        Radiator_order = []
        for key, cb in Filter_radiator_boxes.items():
            if key in active_filter_key_13:
                Radiator_order.append(cb)

        Filter_radiator_box.children = Radiator_order
        update_top_filter_row()
        update_max_parts_slider()


def Filter_heat_valve_updater(change):
    with Filter_heat_valves_output:
        Filter_heat_valves_output.clear_output()
        Selected_items_14 = [key for key, cb in Item_checkboxes.items() if cb.value]
        active_filter_key_14 = set()

        for item_14 in Selected_items_14:
            heat_valve_entry = Checkbox_dictionary_Filter.get(item_14)
            if isinstance(heat_valve_entry, list):
                active_filter_key_14.update(heat_valve_entry)
            elif isinstance(heat_valve_entry, str):
                active_filter_key_14.add(heat_valve_entry)

        Heat_valve_order = []
        for key, cb in Filter_heat_valves_boxes.items():
            if key in active_filter_key_14:
                Heat_valve_order.append(cb)

        Filter_heat_valves_box.children = Heat_valve_order
        update_top_filter_row()
        update_max_parts_slider()


Top_filter_row = widgets.HBox()


#===============================================================================================================================================#
# 9. Dataframe filtering:


def update_top_filter_row():
    general_box = widgets.VBox([widgets.Label("General filters"), Filter_box]) if Filter_box.children else None
    inches_box = widgets.VBox([widgets.Label("Inches diameter"), Filter_inches_box]) if Filter_inches_box.children else None
    sewage_box = widgets.VBox([widgets.Label("Sewage pipe diameter"),\
                               Filter_sewage_pipe_diameter_box]) if Filter_sewage_pipe_diameter_box.children else None
    sewage_clam_box = widgets.VBox([widgets.Label("Pipe clam diameter"),\
                                     Filter_sewage_clam_diameter_box]) if Filter_sewage_clam_diameter_box.children else None
    chrome_box = widgets.VBox([widgets.Label("Chrome diameter"), Filter_chrome_box]) if Filter_chrome_box.children else None
    toilet_box = widgets.VBox([widgets.Label("Wc filters"), Filter_toilet_box]) if Filter_toilet_box.children else None
    heater_box = widgets.VBox([widgets.Label("Parts filters"), Filter_heater_box]) if Filter_heater_box.children else None
    PVC_pipe_box = widgets.VBox([widgets.Label("PVC pipe diameter"), Filter_PVC_pipe_box]) if Filter_PVC_pipe_box.children else None
    Radiator_box = widgets.VBox([widgets.Label("Radiator filters"), Filter_radiator_box]) if Filter_radiator_box.children else None
    Heat_valve_box = widgets.VBox([widgets.Label("Valve filters"), Filter_heat_valves_box]) if Filter_heat_valves_box.children else None
    
    
    pipe_content = []
    if Filter_pipe_box.children:
        pipe_content.append(widgets.Label("Pipe filters"))
        pipe_content.extend(Filter_pipe_box.children)
    if Filter_pipe_diameter_box.children:
        pipe_content.append(widgets.Label("Pipe Diameters"))
        pipe_content.extend(Filter_pipe_diameter_box.children)

    pipe_box = widgets.VBox(pipe_content) if pipe_content else None

    
    press_content = []
    if Filter_press_pipe_box.children:
        press_content.append(widgets.Label("Press filters"))
        press_content.extend(Filter_press_pipe_box.children)
    if Filter_press_diameter_box.children:
        press_content.append(widgets.Label("Press Diameters"))
        press_content.extend(Filter_press_diameter_box.children)

    press_box = widgets.VBox(press_content) if press_content else None

    
    row_content = []
    if general_box: row_content.append(general_box)
    if inches_box: row_content.append(inches_box)
    if pipe_box: row_content.append(pipe_box)
    if sewage_box: row_content.append(sewage_box)
    if sewage_clam_box: row_content.append(sewage_clam_box)
    if chrome_box: row_content.append(chrome_box)
    if toilet_box: row_content.append(toilet_box)
    if heater_box: row_content.append(heater_box)
    if press_box: row_content.append(press_box)
    if PVC_pipe_box: row_content.append(PVC_pipe_box)
    if Radiator_box: row_content.append(Radiator_box)
    if Heat_valve_box: row_content.append(Heat_valve_box)


    Top_filter_row.children = row_content


def Invoice_filtering(change):
    global selected_df
    with Output:
        Output.clear_output()
        
        selected_keys = [key for key, cb in Item_checkboxes.items() if cb.value]
        filtered_dfs = []

        for key in selected_keys:
            if key not in Combined_df:
                continue
            
            df = Combined_df[key].copy()
            mask = pd.Series([True] * len(df), index=df.index)

            
            # ===== 1. PIPE-SPECIFIC FILTERING (Hot/Cold + Diameter) =====
            if key == "Pipes":
                pipe_type_mask = pd.Series([False] * len(df), index=df.index)
                for pipe_key, cb in Filter_pipe_boxes.items():
                    if cb.value:
                        regex = Filter_match.get(pipe_key, "")
                        pipe_type_mask |= df["Name"].str.contains(regex)
                
                if pipe_type_mask.any():
                    mask &= pipe_type_mask

                pipe_diam_mask = pd.Series([False] * len(df), index=df.index)
                for diam_key, cb in Filter_pipe_diameter_boxes.items():
                    if cb.value:
                        regex_pipe = Filter_match.get(diam_key, "")
                        pipe_diam_mask |=df["Name"].str.contains(regex_pipe)
                
                if pipe_diam_mask.any():
                    mask &= pipe_diam_mask

            
            # ===== 2. DIAMETER FILTERING FOR OTHER ITEMS =====
            elif key in ["Plastic fittings", "Plastic reductions", "Modular plastic", "Stock plastic fittings",\
                         "Modular plastic fittings", "Pipe brackets", "PE fittings"]:
                diam_mask = pd.Series([False] * len(df), index=df.index)
                for diam_fitting_key, cb in Filter_boxes.items():
                    if cb.value:
                        regex_diam = Filter_match.get(diam_fitting_key, "")
                        diam_mask |= df["Name"].str.contains(regex_diam)
                        
                if diam_mask.any():
                    mask &= diam_mask

            
            # ===== 3. PIPE CLAMS (Only D25) =====
            elif key == "Pipe clams":
                if Filter_boxes["D25"].value:
                    mask &= df["Name"].str.contains("25")

            
            # ===== 4. Valves  =====
            elif key in ["Valves", "Black metal", "Zinc fittings"]:
                inches_mask = pd.Series([False] * len(df), index=df.index)
                for inch_key, cb in Filter_inches_boxes.items():
                    if cb.value:
                        regex_inches = Filter_match.get(inch_key, "")
                        inches_mask |= df["Name"].str.contains(regex_inches)
                        
                if inches_mask.any():
                    mask &= inches_mask

            
            # ===== 5. Rubber seals  =====
            elif key in ["Rubber seals"]:
                rubber_mask = pd.Series([False] * len(df), index=df.index)
                for rubber_key, cb in Filter_inches_boxes.items():
                    if cb.value:
                        regex_rubber = Filter_match.get(rubber_key, "")
                        rubber_mask |= df["Name"].str.contains(regex_rubber)
                        
                if rubber_mask.any():
                    mask &= rubber_mask


            # ===== 6. Bronze fittings  =====
            elif key in ["Bronze fittings"]:
                bronze_mask = pd.Series([False] * len(df), index=df.index)
                for bronze_key, cb in Filter_inches_boxes.items():
                    if cb.value:
                        regex_bronze = Filter_match.get(bronze_key, "")
                        bronze_mask |= df["Name"].str.contains(regex_bronze)

                if bronze_mask.any():
                    mask &= bronze_mask


            # ===== 7. Sewage pipes  =====
            elif key in ["Sewage pipes"]:
                sewage_mask = pd.Series([False] * len(df), index=df.index)
                for sewage_key, cb in Filter_sewage_pipe_diamter_boxes.items():
                    if cb.value:
                        regex_sewage = Filter_match.get(sewage_key, "")
                        sewage_mask |= df["Name"].str.contains(regex_sewage)

                if sewage_mask.any():
                    mask &= sewage_mask

            
            # ===== 8. Sewage clams  =====
            elif key in ["Sewage clams"]:
                sewage_clam_mask = pd.Series([False] * len(df), index=df.index)
                for sewage_clam_key, cb in Filter_sewage_pipe_clams_boxes.items():
                    if cb.value:
                        regex_sewage_clam = Filter_match.get(sewage_clam_key, "")
                        sewage_clam_mask |= df["Name"].str.contains(regex_sewage_clam)

                if sewage_clam_mask.any():
                    mask &= sewage_clam_mask

            
            # ===== 9. Chrome parts  =====
            elif key in ["Chrome parts", "Extenders"]:
                chrome_mask = pd.Series([False] * len(df), index=df.index)
                for chrome_key, cb in Filter_chrome_boxes.items():
                    if cb.value:
                        regex_chrome = Filter_match.get(chrome_key, "")
                        chrome_mask |= df["Name"].str.contains(regex_chrome)

                if chrome_mask.any():
                    mask &= chrome_mask

            
            # ===== 10. Toilet parts  =====
            elif key in ["Toilets"]:
                toilet_mask = pd.Series([False] * len(df), index=df.index)
                for toilet_key, cb in Filter_toilet_boxes.items():
                    if cb.value:
                        regex_toilet = Filter_match.get(toilet_key, "")
                        toilet_mask |= df.index.str.contains(regex_toilet)

                if toilet_mask.any():
                    mask &= toilet_mask

            
            # ===== 11. Heater parts  =====
            elif key in ["Water heaters"]:
                heater_mask = pd.Series([False] * len(df), index=df.index)
                for heater_key, cb in Filter_heater_boxes.items():
                    if cb.value:
                        regex_heater = Filter_match.get(heater_key, "")
                        heater_mask |= df.index.str.contains(regex_heater)

                if heater_mask.any():
                    mask &= heater_mask

            
            # ===== 12. PRESS-SPECIFIC FILTERING (Multilayer/Metal + Diameter) =====
            elif key in {"Press fittings", "Press pipes"}:
                press_type_mask = pd.Series([False] * len(df), index=df.index)
                for press_key, cb in Filter_press_pipe_boxes.items():
                    if cb.value:
                        regex_press = Filter_match.get(press_key, "")
                        press_type_mask |= df["Name"].str.contains(regex_press)
                
                if press_type_mask.any():
                    mask &= press_type_mask

                press_diam_mask = pd.Series([False] * len(df), index=df.index)
                for press_diam_key, cb in Filter_press_diameter_boxes.items():
                    if cb.value:
                        if "16" in press_diam_key:
                            press_diam_mask |= df["Name"].str.contains("16")
                        elif "18" in press_diam_key:
                            press_diam_mask |= df["Name"].str.contains("18")
                        elif "20" in press_diam_key:
                            press_diam_mask |= df["Name"].str.contains("20")
                        elif "22" in press_diam_key:
                            press_diam_mask |= df["Name"].str.contains("22")
                        elif "26" in press_diam_key:
                            press_diam_mask |= df["Name"].str.contains("26")
                        elif "28" in press_diam_key:
                            press_diam_mask |= df["Name"].str.contains("28")
                        elif "32" in press_diam_key:
                            press_diam_mask |= df["Name"].str.contains("32")
                        elif "35" in press_diam_key:
                            press_diam_mask |= df["Name"].str.contains("35")
                        elif "42" in press_diam_key:
                            press_diam_mask |= df["Name"].str.contains("42")
                            
                if press_diam_mask.any():
                    mask &= press_diam_mask

            
            # ===== 13. PVC pipes  =====
            elif key in ["PVC pipes", "PVC fittings"]:
                PVC_pipe_mask = pd.Series([False] * len(df), index=df.index)
                for PVC_key, cb in Filter_PVC_pipe_boxes.items():
                    if cb.value:
                        regex_PVC = Filter_match.get(PVC_key, "")
                        PVC_pipe_mask |= df["Name"].str.contains(regex_PVC)

                if PVC_pipe_mask.any():
                    mask &= PVC_pipe_mask

            
            # ===== 14. PVC pipes  =====
            elif key in ["Radiators"]:
                Radiator_mask = pd.Series([False] * len(df), index=df.index)
                for Radiator_key, cb in Filter_radiator_boxes.items():
                    if cb.value:
                        regex_radiator = Filter_match.get(Radiator_key, "")
                        Radiator_mask |= df.index.str.contains(regex_radiator)

                if Radiator_mask.any():
                    mask &= Radiator_mask
                    

            # ===== 15. Heating valves  =====
            elif key in ["Heating valves"]:
                Heat_valve_mask = pd.Series([False] * len(df), index=df.index)
                for heat_valve_key, cb in Filter_heat_valves_boxes.items():
                    if cb.value:
                        regex_heat_valve = Filter_match.get(heat_valve_key, "")
                        Heat_valve_mask |= df["Name"].str.contains(regex_heat_valve)

                if Heat_valve_mask.any():
                    mask &= Heat_valve_mask
            

            
            filtered_df = df[mask]
            if not filtered_df.empty:
                filtered_dfs.append(filtered_df)

        if not filtered_dfs:
            print("Please select product group.")
            selected_df = pd.DataFrame()
            Invoice_filtering.last_filtered_df = selected_df
            update_max_parts_slider()
            return
        
        selected_df = pd.concat(filtered_dfs, axis=0)
        Invoice_filtering.last_filtered_df = selected_df
        update_max_parts_slider()
        display(selected_df)


#===============================================================================================================================================#
# 10. Max sliders and filter positions:

def update_group_common_filters(*args):
    group_sliders = []

    if common_items_toggle.value:
        group_common_filter_box.layout.display = 'block'
        selected_groups = [k for k, cb in Item_checkboxes.items() if cb.value]

        for group in selected_groups:
            group_checkbox = widgets.Checkbox(
                description=group,
                value=True,
                layout=widgets.Layout(width="200px", margin="0 0 0 -70px"))

            initial_max = 10  # Default value
            if group in Combined_df:
                df = Combined_df[group].copy()
                mask = pd.Series([True] * len(df), index=df.index)
                
                if group in ["Valves", "Rubber seals", "Bronze fittings", "Black metal"]:
                    inches_mask = pd.Series([False] * len(df), index=df.index)
                    for inch_key, cb in Filter_inches_boxes.items():
                        if cb.value:
                            regex_inches = Filter_match.get(inch_key, "")
                            inches_mask |= df["Name"].str.contains(regex_inches)
                    if inches_mask.any():
                        mask &= inches_mask

                
                elif group in ["Sewage pipes"]:
                    sewage_mask = pd.Series([False] * len(df), index=df.index)
                    for sewage_key, cb in Filter_sewage_pipe_diamter_boxes.items():
                        if cb.value:
                            regex_sewage = Filter_match.get(sewage_key, "")
                            sewage_mask |= df["Name"].str.contains(regex_sewage)
    
                    if sewage_mask.any():
                        mask &= sewage_mask

                
                elif group in ["Sewage clams"]:
                    sewage_clam_mask = pd.Series([False] * len(df), index=df.index)
                    for sewage_clam_key, cb in Filter_sewage_pipe_clams_boxes.items():
                        if cb.value:
                            regex_sewage_clam = Filter_match.get(sewage_clam_key, "")
                            sewage_clam_mask |= df["Name"].str.contains(regex_sewage_clam)
    
                    if sewage_clam_mask.any():
                        mask &= sewage_clam_mask


                elif group in ["Chrome parts", "Extenders"]:
                    chrome_mask = pd.Series([False] * len(df), index=df.index)
                    for chrome_key, cb in Filter_chrome_boxes.items():
                        if cb.value:
                            regex_chrome = Filter_match.get(chrome_key, "")
                            chrome_mask |= df["Name"].str.contains(regex_chrome)
    
                    if chrome_mask.any():
                        mask &= chrome_mask

                        
                elif group in ["Toilets"]:
                    toilet_mask = pd.Series([False] * len(df), index=df.index)
                    for toilet_key, cb in Filter_toilet_boxes.items():
                        if cb.value:
                            regex_toilet = Filter_match.get(toilet_key, "")
                            toilet_mask |= df.index.str.contains(regex_toilet)
    
                    if toilet_mask.any():
                        mask &= toilet_mask

                        
                elif group in ["Water heaters"]:
                    heater_mask = pd.Series([False] * len(df), index=df.index)
                    for heater_key, cb in Filter_heater_boxes.items():
                        if cb.value:
                            regex_heater = Filter_match.get(heater_key, "")
                            heater_mask |= df.index.str.contains(regex_heater)
    
                    if heater_mask.any():
                        mask &= heater_mask

            
                elif group in ["Press fittings", "Press pipes"]:
                    press_type_mask = pd.Series([False] * len(df), index=df.index)
                    for press_key, cb in Filter_press_pipe_boxes.items():
                        if cb.value:
                            regex_press = Filter_match.get(press_key, "")
                            press_type_mask |= df["Name"].str.contains(regex_press)
                    
                    if press_type_mask.any():
                        mask &= press_type_mask
                
                    press_diam_mask = pd.Series([False] * len(df), index=df.index)
                    for press_diam_key, cb in Filter_press_diameter_boxes.items():
                        if cb.value:
                            if "16" in press_diam_key:
                                press_diam_mask |= df["Name"].str.contains("16")
                            elif "18" in press_diam_key:
                                press_diam_mask |= df["Name"].str.contains("18")
                            elif "18" in press_diam_key:
                                press_diam_mask |= df["Name"].str.contains("18")
                            elif "20" in press_diam_key:
                                press_diam_mask |= df["Name"].str.contains("20")
                            elif "22" in press_diam_key:
                                press_diam_mask |= df["Name"].str.contains("22")
                            elif "26" in press_diam_key:
                                press_diam_mask |= df["Name"].str.contains("26")
                            elif "28" in press_diam_key:
                                press_diam_mask |= df["Name"].str.contains("28")
                            elif "32" in press_diam_key:
                                press_diam_mask |= df["Name"].str.contains("32")
                            elif "35" in press_diam_key:
                                press_diam_mask |= df["Name"].str.contains("35")
                            elif "42" in press_diam_key:
                                press_diam_mask |= df["Name"].str.contains("42")                    
                                
                    if press_diam_mask.any():
                        mask &= press_diam_mask


                elif group in ["PVC pipes", "PVC fittings"]:
                    PVC_mask = pd.Series([False] * len(df), index=df.index)
                    for PVC_key, cb in Filter_PVC_pipe_boxes.items():
                        if cb.value:
                            regex_regex_PVC = Filter_match.get(PVC_key, "")
                            PVC_mask |= df["Name"].str.contains(regex_PVC)
    
                    if PVC_mask.any():
                        mask &= PVC_mask
                        

                elif group in ["Radiators"]:
                    Radiator_mask = pd.Series([False] * len(df), index=df.index)
                    for radiator_key, cb in Filter_radiator_boxes.items():
                        if cb.value:
                            regex_radiator = Filter_match.get(radiator_key, "")
                            Radiator_mask |= df.index.str.contains(regex_radiator)
    
                    if Radiator_mask.any():
                        mask &= Radiator_mask
                        

                elif group in ["Heating valves"]:
                    Heat_valves_mask = pd.Series([False] * len(df), index=df.index)
                    for heat_valve_key, cb in Filter_heat_valves_boxes.items():
                        if cb.value:
                            regex_heat_valve = Filter_match.get(heat_valve_key, "")
                            Heat_valves_mask |= df["Name"].str.contains(regex_heat_valve)
    
                    if Heat_valves_mask.any():
                        mask &= Heat_valves_mask

            
            slider = widgets.IntSlider(
                value=min(3, initial_max),
                min=1,
                max=initial_max,
                step=1,
                description='Max parts:',
                layout=widgets.Layout(width="250px"))

            slider_box = widgets.HBox([widgets.Label("→"), slider])
            slider_box.layout.display = 'flex'

            def on_group_checked(change, box=slider_box, group=group):
                box.layout.display = 'flex' if change['new'] else 'none'
                if change['new']:
                    update_max_parts_slider()

            group_checkbox.observe(on_group_checked, 'value')
            group_sliders.append(widgets.HBox([group_checkbox, slider_box]))
    else:
        group_common_filter_box.layout.display = 'none'

    group_common_filter_box.children = group_sliders
    update_max_parts_slider()


def assemble_filter_layout():
    left_half = Top_filter_row
    vertical_line = widgets.HTML(f"""<div style="height: 190px; border-left: 2px solid rgb({UI_color});"></div>""")
    right_half = widgets.VBox([common_items_toggle, group_common_filter_box])
    
    return widgets.HBox([left_half, vertical_line, right_half])


common_mode = common_items_toggle.value


group_limits = {}
if common_mode:
    for widget_box in group_common_filter_box.children:
        checkbox, slider_box = widget_box.children
        if checkbox.value:
            group_name = checkbox.description
            slider = slider_box.children[1]
            group_limits[group_name] = slider.value


for checkbox in Filter_boxes.values():
    checkbox.observe(update_max_parts_slider, 'value')
    
for checkbox in Filter_pipe_boxes.values():
    checkbox.observe(update_max_parts_slider, 'value')
    
for checkbox in Filter_pipe_diameter_boxes.values():
    checkbox.observe(update_max_parts_slider, 'value')
    
for checkbox in Filter_inches_boxes.values():
    checkbox.observe(update_max_parts_slider, 'value')

for checkbox in Filter_sewage_pipe_diamter_boxes.values():
    checkbox.observe(update_max_parts_slider, 'value')

for checkbox in Filter_sewage_pipe_clams_boxes.values():
    checkbox.observe(update_max_parts_slider, 'value')

for checkbox in Filter_chrome_boxes.values():
    checkbox.observe(update_max_parts_slider, 'value')
    
for checkbox in Filter_toilet_boxes.values():
    checkbox.observe(update_max_parts_slider, 'value')

for checkbox in Filter_heater_boxes.values():
    checkbox.observe(update_max_parts_slider, 'value')

for checkbox in Filter_press_pipe_boxes.values():
    checkbox.observe(update_max_parts_slider, 'value')

for checkbox in Filter_press_diameter_boxes.values():
    checkbox.observe(update_max_parts_slider, 'value')

for checkbox in Filter_PVC_pipe_boxes.values():
    checkbox.observe(update_max_parts_slider, 'value')

for checkbox in Filter_radiator_boxes.values():
    checkbox.observe(update_max_parts_slider, 'value')

for checkbox in Filter_heat_valves_boxes.values():
    checkbox.observe(update_max_parts_slider, 'value')
    
common_items_toggle.observe(update_group_common_filters, 'value')
group_common_filter_box = widgets.VBox()
group_common_filter_box.layout.display = 'none'


#===============================================================================================================================================#
# 11. Invoice settings:

Price_text = widgets.Text(description='Invoice price:', layout=widgets.Layout(margin='0 0 0 0', width='200px'))
Price_upper_limit = widgets.Text(description="Price upper limit:", layout=widgets.Layout(margin='0 0 0 0', width='250px'),\
                               style={'description_width': '120px'})
Price_lower_limit = widgets.Text(description="Price lower limit:", layout=widgets.Layout(margin='0 0 0 0', width='250px'),\
                               style={'description_width': '120px'})

Invoice_price_settings = widgets.HBox([Price_text, Price_lower_limit, Price_upper_limit]) 

Price_value = None
Lower_limit_value = None
Upper_limit_value = None

def on_price_input_change(change):
    global Price_value
    try:
        Price_value = float(change["new"])
    except ValueError:
        Price_value = None

def on_lower_limit_change(change):
    global Lower_limit_value
    try:
        Lower_limit_value = float(change["new"])
    except ValueError:
        Lower_limit_value = None

def on_upper_limit_change(change):
    global Upper_limit_value
    try:
        Upper_limit_value = float(change["new"])
    except ValueError:
        Upper_limit_value = None


###############################################################
# Invoice_generator bellow is used to set bounds for upper and lower amount
# for the invoices, hence its needs to be updated just like the code above
# without it, the groups that are in the code, will only be affected by upper
# and lower bounds that you set in the text fields. Meaning if you set between
# 50-100, invoice will be in that range, otherwise it will ignore it.
###############################################################


def Invoice_generator(change):
    global Invoice
    with Output:
        Output.clear_output()
        missing_fields = []

        if Price_value is None:
            missing_fields.append("Invoice price")
        if Lower_limit_value is None:
            missing_fields.append("Price lower limit")
        if Upper_limit_value is None:
            missing_fields.append("Price upper limit")

        if missing_fields:
            print("Missing or invalid input:", ", ".join(missing_fields))
            return

        if common_items_toggle.value:
            group_limits = {}
            for widget_box in group_common_filter_box.children:
                checkbox, slider_box = widget_box.children
                if checkbox.value:
                    group_name = checkbox.description
                    slider = slider_box.children[1]
                    group_limits[group_name] = slider.value

            filtered_dfs = []
            for group_name, max_parts in group_limits.items():
                if group_name in Combined_df:
                    group_df = Combined_df[group_name].copy()
                    mask = pd.Series([True] * len(group_df), index=group_df.index)

                    if group_name in ["Valves", "Rubber seals", "Bronze fittings", "Black metal", "Zinc fittings"]:
                        inches_mask = pd.Series([False] * len(group_df), index=group_df.index)
                        for inch_key, cb in Filter_inches_boxes.items():
                            if cb.value:
                                regex_inches = Filter_match.get(inch_key, "")
                                inches_mask |= group_df["Name"].str.contains(regex_inches)
                        if inches_mask.any():
                            mask &= inches_mask
                  
                    elif group_name in ["Sewage pipes"]:
                        sewage_mask = pd.Series([False] * len(group_df), index=group_df.index)
                        for sewage_diameter, cb in Filter_sewage_pipe_diamter_boxes.items():
                            if cb.value:
                                regex_sewage = Filter_match.get(sewage_diameter, "")
                                sewage_mask |= group_df["Name"].str.contains(regex_sewage)
                        if sewage_mask.any():
                            mask &= sewage_mask
                    
                    elif group_name in ["Sewage clams"]:
                        sewage_clam_mask = pd.Series([False] * len(group_df), index=group_df.index)
                        for clam_diameter, cb in Filter_sewage_pipe_clams_boxes.items():
                            if cb.value:
                                regex_sewage_clam = Filter_match.get(clam_diameter, "")
                                sewage_clam_mask |= group_df["Name"].str.contains(regex_sewage_clam)
                        if sewage_clam_mask.any():
                            mask &= sewage_clam_mask

                    elif group_name in ["Chrome parts", "Extenders"]:
                        chrome_mask = pd.Series([False] * len(group_df), index=group_df.index)
                        for chrome_diameter, cb in Filter_chrome_boxes.items():
                            if cb.value:
                                regex_chrome = Filter_match.get(chrome_diameter, "")
                                chrome_mask |= group_df["Name"].str.contains(regex_chrome)
                        if chrome_mask.any():
                            mask &= chrome_mask

                        
                    elif group_name in ["Toilets"]:
                        toilet_mask = pd.Series([False] * len(group_df), index=group_df.index)
                        for toilet_check, cb in Filter_toilet_boxes.items():
                            if cb.value:
                                regex_toilet = Filter_match.get(toilet_check, "")
                                toilet_mask |= group_df.index.str.contains(regex_toilet)
                        if toilet_mask.any():
                            mask &= toilet_mask
                              
                        
                    elif group_name in ["Water heaters"]:
                        heater_mask = pd.Series([False] * len(group_df), index=group_df.index)
                        for heater_check, cb in Filter_heater_boxes.items():
                            if cb.value:
                                regex_heater = Filter_match.get(heater_check, "")
                                heater_mask |= group_df.index.str.contains(regex_heater)
                        if heater_mask.any():
                            mask &= heater_mask
                            
                            
                    elif group_name in ["Pipes", "Plastic fittings", "Plastic reductions", "Modular plastic",\
                                        "Stock plastic fittings", "Modular plastic fittings", "Pipe brackets",\
                                        "PE fittings"]:
                        diam_mask = pd.Series([False] * len(group_df), index=group_df.index)
                        if group_name == "Pipes":
                            for diam_key, cb in Filter_pipe_diameter_boxes.items():
                                if cb.value:
                                    regex_pipe = Filter_match.get(diam_key, "")
                                    diam_mask |= df["Name"].str.contains(regex_pipe)
                        else:
                            for fitting_key, cb in Filter_boxes.items():
                                if cb.value:
                                    regex_diam = Filter_match.get(fitting_key, "")
                                    diam_mask |= group_df["Name"].str.contains(regex_diam)
                        if diam_mask.any():
                            mask &= diam_mask

                    
                    elif group_name in ["Press fittings", "Press pipes"]: 
                        Press_mask = pd.Series([False] * len(group_df), index=group_df.index)
                        for press_check, cb in Filter_press_diameter_boxes.items():
                            if cb.value:
                                regex_press = Filter_match.get(press_check, "")
                                Press_mask |= group_df["Name"].str.contains(regex_press)
                        if Press_mask.any():
                            mask &= Press_mask

                            
                    elif group_name in ["PVC pipes", "PVC fittings"]: 
                        PVC_mask = pd.Series([False] * len(group_df), index=group_df.index)
                        for PVC_check, cb in Filter_PVC_pipe_boxes.items():
                            if cb.value:
                                regex_PVC = Filter_match.get(PVC_check, "")
                                PVC_mask |= group_df["Name"].str.contains(regex_PVC)
                        if PVC_mask.any():
                            mask &= PVC_mask
                            
                            
                    elif group_name in ["Radiators"]: 
                        Radiator_mask = pd.Series([False] * len(group_df), index=group_df.index)
                        for radiator_check, cb in Filter_radiator_boxes.items():
                            if cb.value:
                                regex_radiator = Filter_match.get(radiator_check, "")
                                Radiator_mask |= group_df.index.str.contains(regex_radiator)
                        if Radiator_mask.any():
                            mask &= Radiator_mask
                            
                            
                    elif group_name in ["Heating valves"]: 
                        heat_valve_mask = pd.Series([False] * len(group_df), index=group_df.index)
                        for heat_valve_check, cb in Filter_heat_valves_boxes.items():
                            if cb.value:
                                regex_heat_valve = Filter_match.get(heat_valve_check, "")
                                heat_valve_mask |= group_df["Name"].str.contains(regex_heat_valve)
                        if heat_valve_mask.any():
                            mask &= heat_valve_mask

                    
                    filtered_group = group_df[mask]
                    if not filtered_group.empty:
                        filtered_dfs.append(filtered_group.head(max_parts))
            
            if filtered_dfs:
                selected_df = pd.concat(filtered_dfs, axis=0)
                Invoice = Invoice_assembler(selected_df, mode='common', max_parts=None)
            else:
                print("No matching items found in selected groups.")
                return
        else:
            if hasattr(Invoice_filtering, 'last_filtered_df') and not Invoice_filtering.last_filtered_df.empty:
                selected_df = Invoice_filtering.last_filtered_df
                Invoice = simple_Invoice_assembler(selected_df)
            else:
                print("No filtered data available.")
                return

        display(Invoice)


Price_text.observe(on_price_input_change, names="value")
Price_lower_limit.observe(on_lower_limit_change, names="value")
Price_upper_limit.observe(on_upper_limit_change, names="value")


#===============================================================================================================================================#
# 12. Client info:

Default = {"Pirkėjas": '',
           "Adresas": "",
           "Įmonės kodas": "",
           "PVM kodas": ""}

Company_A = {"Pirkėjas": 'Company A',
           "Adresas": "Lithuania",
           "Įmonės kodas": "11111111111",
           "PVM kodas": "LT11111111111"}

Company_B = {"Pirkėjas": 'Company B',
           "Adresas": "Lithuania",
           "Įmonės kodas": "222222222222",
           "PVM kodas": "LT222222222222"}


Client_menu = widgets.Dropdown(options = [("Company A", Company_A),
                                          ("Company B", Company_B),
                                          ("Empty", Default)], 
                               value = Default, description = "Client menu", layout=widgets.Layout(margin='0 0 0 0', width='300px'))


Client = widgets.Text(description='Company name:', layout=widgets.Layout(margin='0 0 0 -25px', width='300px'),\
                      style={'description_width': '120px'}, continuous_update=True)

Address = widgets.Text(description="Address:", layout=widgets.Layout(margin='0 0 0 -50px', width='400px'),\
                       style={'description_width': '120px'}, continuous_update=True)

Company_code = widgets.Text(description="Company code:", layout=widgets.Layout(margin='0 0 0 0', width='250px'),\
                            style={'description_width': '120px'}, continuous_update=True)

VAT_code = widgets.Text(description="VAT code:", layout=widgets.Layout(margin= "0 0 0 -50px", width="250px"),\
                        style={'description_width': '120px'}, continuous_update=True)

Client_info_box = widgets.HBox([Client, Address, Company_code, VAT_code], layout=widgets.Layout(margin="20px 0 0 0"))

 
Client_info_box.layout.display = 'flex' if Client_menu.value == Default else 'none'
def toggle_client_info(change):
    if change['name'] == 'value':
        if change['new'] == Default:
            Client_info_box.layout.display = 'flex'
            update_client_cells(Default)
            
        else:
            Client_info_box.layout.display = 'none'
            update_client_cells(change['new'])

Client_menu.observe(toggle_client_info, names='value')

Invoice_name_text = widgets.Text(description="Invoice name:", layout=widgets.Layout(margin='0 0 0 0', width='300px'))
Invoice_text_box = widgets.HBox([Invoice_name_text, Client_menu])


def update_client_cells(client_data):
    """Updates the client information cells in the Excel worksheet"""
    WS["F6"].value = client_data["Pirkėjas"]  # Company name
    WS["H7"].value = client_data["Adresas"]   # Address
    WS["H8"].value = client_data["Įmonės kodas"]  # Company code
    WS["H9"].value = client_data["PVM kodas"]  # VAT code


def handle_client_change(change):
    if change['name'] == 'value':
        selected_client = change['new']
        update_client_cells(selected_client)
        
        Client.value = selected_client["Pirkėjas"]
        Address.value = selected_client["Adresas"]
        Company_code.value = str(selected_client["Įmonės kodas"])
        VAT_code.value = selected_client["PVM kodas"]


def handle_text_change(change):
    if Client_menu.value == Default:  
        updated_client = {"Pirkėjas": Client.value,
                          "Adresas": Address.value,
                          "Įmonės kodas": Company_code.value,
                          "PVM kodas": VAT_code.value}
        update_client_cells(updated_client)


Client_menu.observe(handle_client_change, names='value')
Client.observe(handle_text_change, names='value')
Address.observe(handle_text_change, names='value')
Company_code.observe(handle_text_change, names='value')
VAT_code.observe(handle_text_change, names='value')


#===============================================================================================================================================#
# 13. Invoice excel formating:

def create_excel_template(Invoice_excel, invoice_number):
    global WS
    WB = Workbook()
    WS = WB.active
    WS.title = f"LES {invoice_number}"
    
    Logo = Image(Logo_path)

    def cm(margin_input):
        return margin_input/ 2.53
    
    WS.page_margins = PageMargins(
        left = cm(0.8),
        right = cm(0.3),
        top = cm(0.9),
        bottom = cm(1.4),
        header = cm(1.3),
        footer = cm(0.8))
    
    WS.page_setup.paperSize = WS.PAPERSIZE_A4
    
    Color_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    
    # 1 column width unit ≈ 7.5 pixels
    # So if we need 80 pixeles, we need to 80 / 7.5 = 10.67, because python and excel measure the same case width differently.
    
    Const = 1.09617755856967
    
    ## Column parameters:
    WS.column_dimensions["A"].width = 8.11 * Const
    WS.column_dimensions["B"].width = 4.18 * Const
    WS.column_dimensions["C"].width = 9.67 * Const
    
    for col in  ["D", "E", "F"]:
        WS.column_dimensions[col].width = 6.77 * Const
        
    WS.column_dimensions["G"].width = 5.98 * Const
    WS.column_dimensions["H"].width = 8.67 * Const
    WS.column_dimensions["I"].width = 7.67 * Const
    WS.column_dimensions["J"].width = 7.32 * Const
    WS.column_dimensions["K"].width = 5.52 * Const
    WS.column_dimensions["L"].width = 8.33 * Const
    WS.column_dimensions["M"].width = 1.54 * Const
    
    
    ## Rows parameters:
    WS.row_dimensions[1].height = 14.40
    WS.row_dimensions[2].height = 21.00
    WS.row_dimensions[3].height = 14.40
    WS.row_dimensions[4].height = 14.40
    WS.row_dimensions[5].height = 18.00
    
    for i in range(17, 60):
        WS.row_dimensions[i].height = 15.00
    
    ## Logo:
    
    Logo.width= 6.90 * 37.8
    Logo.height = 1.32 * 37.8
    WS.add_image(Logo, "A2")
    
    ## Text:
    
    WS.merge_cells("F2:I2")
    Cell_1 = WS["F2"]
    Cell_1.value = "PVM SĄSKAITA FAKTŪRA"
    Cell_1.font = Font(name="Calibri", size=16, bold=True)
    Cell_1.alignment = Alignment(horizontal="right", vertical="bottom")
    
    Cell_2 = WS["J2"]
    Cell_2.value = "Nr."
    Cell_2.font = Font(name="Calibri", size=16, bold=True)
    Cell_2.alignment = Alignment(horizontal="right", vertical="bottom")
    
    Cell_3 = WS["K2"]
    Cell_3.value = f"LES {invoice_number}"
    Cell_3.font = Font(name="Calibri", size=16, bold=True)
    Cell_3.alignment = Alignment(vertical="bottom")
    
    WS.merge_cells("A5:B5")
    Cell_4 = WS["A5"]
    Cell_4.value = "Pardavėjas:"
    Cell_4.font = Font(name="Calibri", size=14, bold=True)
    Cell_4.alignment = Alignment(horizontal="left", vertical="bottom")
    
    Cell_5 = WS["A6"]
    Cell_5.value = 'My company'
    Cell_5.font = Font(name="Calibri", size=11, bold=True)
    Cell_5.alignment = Alignment(vertical="bottom")
    
    Cell_6 = WS["A7"]
    Cell_6.value = "Adresas:"
    Cell_6.font = Font(name="Calibri", size=11)
    Cell_6.alignment = Alignment(vertical="bottom")
    
    Cell_7 = WS["A8"]
    Cell_7.value = "Įmonės kodas:"
    Cell_7.font = Font(name="Calibri", size=11)
    Cell_7.alignment = Alignment(vertical="bottom")
    
    Cell_8 = WS["A9"]
    Cell_8.value = "PVM kodas:"
    Cell_8.font = Font(name="Calibri", size=11)
    Cell_8.alignment = Alignment(vertical="bottom")
    
    Cell_9 = WS["A10"]
    Cell_9.value = "Banko kodas:"
    Cell_9.font = Font(name="Calibri", size=11)
    Cell_9.alignment = Alignment(vertical="bottom")
    
    Cell_10 = WS["A10"]
    Cell_10.value = "Banko kodas:"
    Cell_10.font = Font(name="Calibri", size=11)
    Cell_10.alignment = Alignment(vertical="bottom")
    
    Cell_11 = WS["A11"]
    Cell_11.value = "A.S."
    Cell_11.font = Font(name="Calibri", size=11)
    Cell_11.alignment = Alignment(vertical="bottom")
    
    Cell_12 = WS["C7"]
    Cell_12.value = "My address"
    Cell_12.font = Font(name="Calibri", size=11)
    Cell_12.alignment = Alignment(vertical="bottom")
    
    Cell_13 = WS["C8"]
    Cell_13.value = "12345"
    Cell_13.font = Font(name="Calibri", size=11)
    Cell_13.alignment = Alignment(vertical="bottom")
    
    Cell_14 = WS["C9"]
    Cell_14.value = "LT12345"
    Cell_14.font = Font(name="Calibri", size=11)
    Cell_14.alignment = Alignment(vertical="bottom")
    
    Cell_15 = WS["C10"]
    Cell_15.value = 'Lithuanian Bank'
    Cell_15.font = Font(name="Calibri", size=11)
    Cell_15.alignment = Alignment(vertical="bottom")
    
    Cell_16 = WS["C11"]
    Cell_16.value = 'LT12345'
    Cell_16.font = Font(name="Calibri", size=11)
    Cell_16.alignment = Alignment(vertical="bottom")
    
    
    ## Pirkejas:
    
    Cell_17 = WS["F5"]
    Cell_17.value = "Pirkėjas:"
    Cell_17.font = Font(name="Calibri", size=14, bold=True)
    Cell_17.alignment = Alignment(horizontal="left", vertical="bottom")
    
    Cell_18 = WS["F6"]
    Cell_18.value = "TEST 1"
    Cell_18.font = Font(name="Calibri", size=11, bold=True)
    Cell_18.alignment = Alignment(vertical="bottom")
    
    Cell_19 = WS["F7"]
    Cell_19.value = "Adresas:"
    Cell_19.font = Font(name="Calibri", size=11)
    Cell_19.alignment = Alignment(vertical="bottom")
    
    Cell_20 = WS["F8"]
    Cell_20.value = "Įmonės kodas:"
    Cell_20.font = Font(name="Calibri", size=11)
    Cell_20.alignment = Alignment(vertical="bottom")
    
    Cell_21 = WS["F9"]
    Cell_21.value = "PVM kodas:"
    Cell_21.font = Font(name="Calibri", size=11)
    Cell_21.alignment = Alignment(vertical="bottom")
    
    Cell_22 = WS["F10"]
    Cell_22.value = "Pastaba:"
    Cell_22.font = Font(name="Calibri", size=11)
    Cell_22.alignment = Alignment(vertical="bottom")
    
    Cell_23 = WS["H7"]
    Cell_23.value = "TEST 2"
    Cell_23.font = Font(name="Calibri", size=11)
    Cell_23.alignment = Alignment(vertical="bottom")
    
    Cell_24 = WS["H8"]
    Cell_24.value = "TEST 3"
    Cell_24.font = Font(name="Calibri", size=11)
    Cell_24.alignment = Alignment(horizontal="left", vertical="bottom")
    
    Cell_25 = WS["H9"]
    Cell_25.value = "TEST 4"
    Cell_25.font = Font(name="Calibri", size=11)
    Cell_25.alignment = Alignment(vertical="bottom")
    
    ## Rest of the document:
    
    WS.merge_cells("I13:K13")
    Cell_26 = WS["I13"]
    Cell_26.value = "Dokumento data:"
    Cell_26.font = Font(name="Calibri", size=11)
    Cell_26.alignment = Alignment(horizontal="center", vertical="bottom")
    
    WS.merge_cells("I14:K14")
    Cell_27 = WS["I14"]
    Cell_27.value = "         Apmokėti iki:                "
    Cell_27.font = Font(name="Calibri", size=11)
    Cell_27.alignment = Alignment(horizontal="center", vertical="bottom")
    
    WS.merge_cells("L13:M13")
    Cell_28 = WS["L13"]
    Cell_28.value = "=TODAY()"
    Cell_28.font = Font(name="Calibri", size=11)
    Cell_28.alignment = Alignment(horizontal="center", vertical="bottom")
    Cell_28 = WS["L13"].number_format = "yyyy-mm-dd"
    
    WS.merge_cells("L14:M14")
    Cell_29 = WS["L14"]
    Cell_29.value = "=L13+31"
    Cell_29.font = Font(name="Calibri", size=11)
    Cell_29.alignment = Alignment(horizontal="center", vertical="bottom")
    Cell_29 = WS["L14"].number_format = "yyyy-mm-dd"
    
    Cell_30 = WS["A15"]
    Cell_30.value = "Kodas"
    Cell_30.font = Font(name="Calibri", size=9, bold=True)
    Cell_30.alignment = Alignment(horizontal="left", vertical="top")
    Cell_30.fill = Color_fill
    
    WS.merge_cells("B15:C15")
    Cell_31 = WS["B15"]
    Cell_31.value = "Pavadinimas"
    Cell_31.font = Font(name="Calibri", size=9, bold=True)
    Cell_31.alignment = Alignment(horizontal="left", vertical="top")
    Cell_31.fill = Color_fill
    
    WS.merge_cells("D15:E15")
    Cell_32 = WS["D15"]
    Cell_32.fill = Color_fill
    
    Cell_33 = WS["F15"]
    Cell_33.value = "Mato vnt."
    Cell_33.font = Font(name="Calibri", size=9, bold=True)
    Cell_33.alignment = Alignment(vertical="top")
    Cell_33.fill = Color_fill
    
    Cell_34 = WS["G15"]
    Cell_34.value = "Kiekis"
    Cell_34.font = Font(name="Calibri", size=9, bold=True)
    Cell_34.alignment = Alignment(horizontal="center", vertical="top")
    Cell_34.fill = Color_fill
    
    WS.merge_cells("H15:H16")
    Cell_35 = WS["H15"]
    Cell_35.value = "Kaina Eur.  (Be PVM)"
    Cell_35.font = Font(name="Calibri", size=9, bold=True)
    Cell_35.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    Cell_35.fill = Color_fill
    
    Cell_36 = WS["I15"]
    Cell_36.value = "PVM %"
    Cell_36.font = Font(name="Calibri", size=9, bold=True)
    Cell_36.alignment = Alignment(horizontal="center", vertical="top")
    Cell_36.fill = Color_fill
    
    WS.merge_cells("J15:K15")
    Cell_37 = WS["J15"]
    Cell_37.value = "PVM Suma Eur."
    Cell_37.font = Font(name="Calibri", size=9, bold=True)
    Cell_37.alignment = Alignment(horizontal="center", vertical="top")
    Cell_37.fill = Color_fill
    
    WS.merge_cells("L15:M16")
    Cell_38 = WS["L15"]
    Cell_38.value = "Suma Eur. (Be PVM)"
    Cell_38.font = Font(name="Calibri", size=9, bold=True)
    Cell_38.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    Cell_38.fill = Color_fill
    
    WS.merge_cells("A16:G16")
    Cell_39 = WS["A16"]
    Cell_39.fill = Color_fill
    
    WS.merge_cells("I16:K16")
    Cell_40 = WS["I16"]
    Cell_40.fill = Color_fill

    ## Invoice Cells:
    
    #                        0    1    2    3    4    5    6    7    8    9    10   11   12
    Moving_cells_columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]

    Invoice_df = Invoice_excel
    global Iteration_min, Iteration_max, Actuall_cells
    Actuall_cells = Invoice_df.shape[0] - 1
    Iteration_min = 17
    Iteration_max = Iteration_min + Invoice_df.shape[0]

    def round_up(value, decimal):
        multiplier = 10 ** decimal
        return math.ceil(value * multiplier) / multiplier
    
    for i in range(Iteration_min, Iteration_max):
        J_merge = f'{str(Moving_cells_columns[9])}{str(i)}'
        K_merge = f'{str(Moving_cells_columns[10])}{str(i)}'
        Merge_JK = f'{J_merge}:{K_merge}'

        L_merge = f'{str(Moving_cells_columns[11])}{str(i)}'
        M_merge = f'{str(Moving_cells_columns[12])}{str(i)}'
        Merge_LM = f'{L_merge}:{M_merge}'

        WS.merge_cells(Merge_JK)
        WS.merge_cells(Merge_LM)

        Cell_columns = [f'{col}{i}' for col in Moving_cells_columns]
        for col in Cell_columns:
            Invoice_cells = WS[col]
            Invoice_cells.border = Border(bottom=Side(style="thin", color="D9D9D9"))

        Df_rows = range(Iteration_min, Iteration_max)
        Temporary_df = pd.DataFrame(index=Df_rows, columns=Moving_cells_columns)

        Border_style = Border(bottom=Side(style="thin", color="D9D9D9"))
        
        New_letters = ["A", "B", "F", "G", "H", "I", "J", "L"]
        Font_style = Font(name="Calibri", size=8)
        
        for df_idx, excel_row in zip(Invoice_df.index, range(Iteration_min, Iteration_max)):
            WS[f"A{excel_row}"].value = df_idx
            WS[f"B{excel_row}"].value = Invoice_df.loc[df_idx, "Name"]
            WS[f"F{excel_row}"].value = "Vnt"
            WS[f"G{excel_row}"].value = Invoice_df.loc[df_idx, "Quantity"]


            Profit = Invoice_df.loc[df_idx, "Dealer_price"] * (Invoice_df.loc[df_idx, "Market_proc"] / 100)
            Unit_price = Invoice_df.loc[df_idx, "Dealer_price"] + Profit
            
            WS[f"H{excel_row}"].value = Unit_price
            WS[f"H{excel_row}"].number_format = '#,##0.00'

            H_cell = f"H{excel_row}"
            G_cell = f"G{excel_row}"
            Tax_sum = f"=ROUND(({H_cell}*{G_cell})*0.21, 2)"
          
            WS[f"I{excel_row}"].value = "21%"
            
            WS[f"J{excel_row}"].value = Tax_sum
            WS[f"J{excel_row}"].number_format = '#,##0.00'

            Final_sum = f'={H_cell}*{G_cell}'
            
            WS[f"L{excel_row}"].value = Final_sum
            WS[f"L{excel_row}"].number_format = '#,##0.00'

            for i in New_letters:
                WS[f'{i}{excel_row}'].font = Font_style
            

            WS[f"A{excel_row}"].alignment = Alignment(horizontal="left", vertical="bottom")
            WS[f"B{excel_row}"].alignment = Alignment(vertical="bottom")
            WS[f"F{excel_row}"].alignment = Alignment(horizontal="center", vertical="bottom")
            WS[f"G{excel_row}"].alignment = Alignment(horizontal="center", vertical="bottom")
            WS[f"H{excel_row}"].alignment = Alignment(horizontal="right", vertical="bottom")
            WS[f"I{excel_row}"].alignment = Alignment(horizontal="right", vertical="bottom")
            WS[f"J{excel_row}"].alignment = Alignment(horizontal="right", vertical="bottom")
            WS[f"L{excel_row}"].alignment = Alignment(horizontal="right", vertical="bottom")

            Cell_sum_no_tax = WS[str(Moving_cells_columns[11]) + str(18 + Actuall_cells)]
            Sum_start_1 = f'L{Iteration_min}'
            Sum_end_1 = f'M{Iteration_max - 1}'
            SUM_1 = f'=SUM({Sum_start_1}:{Sum_end_1})'
            Cell_sum_no_tax.value = SUM_1
            Cell_sum_no_tax.font = Font(name="Calibri", size=9)
            Cell_sum_no_tax.alignment = Alignment(horizontal="right", vertical="bottom")
            Cell_sum_no_tax.number_format = '#,##0.00'
            Cell_sum_no_tax.border = Border(bottom=Side(style="thin", color="D9D9D9"))

            Cell_vat_sum = WS[str(Moving_cells_columns[11]) + str(20 + Actuall_cells)]
            Sum_vat_start = f'J{Iteration_min}'
            Sum_vat_end = f'K{Iteration_max - 1}'
            Vat_sum = f'=SUM({Sum_vat_start}:{Sum_vat_end})'
            Cell_vat_sum.value = Vat_sum
            Cell_vat_sum.font = Font(name="Calibri", size=9)
            Cell_vat_sum.alignment = Alignment(horizontal="right", vertical="bottom")
            Cell_vat_sum.number_format = '#,##0.00'
            Cell_vat_sum.border = Border(bottom=Side(style="thin", color="D9D9D9"))

            total_no_vat = 0.0
            total_vat = 0.0
            
            for df_idx, excel_row in zip(Invoice_df.index, range(Iteration_min, Iteration_max)):
        
                
                Profit_float = Invoice_df.loc[df_idx, "Dealer_price"] * (Invoice_df.loc[df_idx, "Market_proc"] / 100)
                Unit_price_float = Invoice_df.loc[df_idx, "Dealer_price"] + Profit_float
                Quantity_float = Invoice_df.loc[df_idx, "Quantity"]
            
                line_total = Unit_price_float * Quantity_float
                vat_amount = line_total * 0.21
            
                total_no_vat += line_total
                total_vat += vat_amount
                
            final_total = round(total_no_vat + total_vat, 2)

        
        WS.merge_cells(f"J{excel_row}:K{excel_row}")
        WS.merge_cells(f"L{excel_row}:M{excel_row}")
    
        for col_letter in Moving_cells_columns:
            WS[f"{col_letter}{excel_row}"].border = Border_style
            

    ## Sum in words:

    Num_word_dictionary = {"0": "null",
                           "1": "vien",
                           "2": "d",
                           "3": "tr",
                           "4": "keturi",
                           "5": "penki",
                           "6": "šeši",
                           "7": "septyni",
                           "8": "aštuoni",
                           "9": "devyni",
                           "10": "dešimt"}
    
    Word_endings_1 = {"null": " eurų",
                      "vien": "as",
                      "d": "u",
                      "tr": "ys",
                      "keturi": "",
                      "penki": "",
                      "šeši": "",
                      "septyni": "",
                      "aštuoni": "",
                      "devyni": "",}
    
    Word_endings_2 = {"10": "dešimt",
                      "11": "vienuolika",
                      "12": "dvylika",
                      "13": "trilika",
                      "14": "keturiolika",
                      "15": "penkiolika",
                      "16": "šešiolika",
                      "17": "septyniolika",
                      "18": "aštuoniolika",
                      "19": "devyniolika",}
    
    Word_endings_3 = {"null": "",
                      "d": "videšimt",
                      "tr": "isdešimt",
                      "keturi": "asdešimt",
                      "penki": "asdešimt",
                      "šeši": "asdešimt",
                      "septyni": "asdešimt",
                      "aštuoni": "asdešimt",
                      "devyni": "asdešimt",}
    
    
    def Sum_to_word_converted(num):
        Split_num = str(num).split(".")
        Length = len(Split_num[0])
    
        Word_list = []
        for i in Split_num[0]:
            for key, word in Num_word_dictionary.items():
                if str(i) == key:
                    Word_list.append(word)
    
        Completed_list_1 = []
    
        if Length == 4:
            for index, word in enumerate(Word_list):
                if index == 0 and word in Word_endings_1:
                    combined = word + Word_endings_1[word]
                    if combined == "vienas":
                        Completed_list_1.append(combined + " tūkstantis")
                    else:
                        Completed_list_1.append(combined + " tūkstančiai")
    
                elif index == 1 and word == "null":
                    Completed_list_1.append("")
                    
                elif index == 1 and word in Word_endings_1:
                    combined = word + Word_endings_1[word]
                    if combined == "vienas":
                        Completed_list_1.append(" " + combined + " šimtas")
                    else:
                        Completed_list_1.append(" " + combined + " šimtai")
    
                elif index == 2 and word == "vien":
                    Num_seperator = str(int(Split_num[0]) - (int(Split_num[0]) // 100) * 100)
                    if Num_seperator in Word_endings_2:
                        middle_number = Word_endings_2[Num_seperator] + " eurų"
                        Completed_list_1.append(" " + middle_number)
    
                elif index == 2 and word == "null":
                    Completed_list_1.append("")
    
                elif index == 2 and word != "vien":
                    combined_2 = word + Word_endings_3[word]
                    if Split_num[0][1] == "0":
                        Completed_list_1.append(" " + combined_2 + " eurų")
                    else:
                        Completed_list_1.append(" " + combined_2)
    
                elif index == 2 and Split_num[0][1] == "1":
                    Completed_list_1.append("")
    
                elif index == 3 and word == "null":
                    Null_val = Word_endings_1[word]
                    Null_val.split(' ')
                    Selected_value = Null_val[0]
                    Completed_list_1.append(" " + Selected_value + " eurų")
    
                elif index == 3 and word != 'null':
                    combined_3 = word + Word_endings_1[word]
                    if combined_3 == "vienas":
                        Completed_list_1.append(" " + combined_3 + " euras")
                    else:
                        Completed_list_1.append(" " + combined_3 + " eurai")
                else:
                    Completed_list_1.append("ERROR")
    
    
        if Length == 3:
            for index, word in enumerate(Word_list):
                if index == 0 and word in Word_endings_1:
                    combined = word + Word_endings_1[word]
                    if combined == "vienas":
                        Completed_list_1.append(combined + " šimtas")
                    else:
                        Completed_list_1.append(combined + " šimtai")
    
                elif index == 1 and word == "vien":
                    Num_seperator = str(int(Split_num[0]) - (int(Split_num[0]) // 100) * 100)
                    if Num_seperator in Word_endings_2:
                        middle_number = Word_endings_2[Num_seperator] + " eurų"
                        Completed_list_1.append(" " + middle_number)
                    else:
                        Completed_list_1.append("error")
    
                elif index == 1 and word == "null":
                    Completed_list_1.append("")
    
                elif index == 1 and word != "vien":
                    combined_2 = word + Word_endings_3[word]
                    if Split_num[0][1] == "0":
                        Completed_list_1.append(" " + combined_2 + " eurų")
                    else:
                        Completed_list_1.append(" " + combined_2)
    
                elif index == 1 and Split_num[0][1] == "1":
                    Completed_list_1.append("")
    
                elif index == 2 and word == "null":
                    Null_val = Word_endings_1[word]
                    Null_val.split(' ')
                    Selected_value = Null_val[0]
                    Completed_list_1.append(" " + Selected_value + " eurų")
    
                elif index == 2 and word != 'null':
                    combined_3 = word + Word_endings_1[word]
                    if combined_3 == "vienas":
                        Completed_list_1.append(" " + combined_3 + " euras")
                    else:
                        Completed_list_1.append(" " + combined_3 + " eurai")
    
                else:
                    Completed_list_1.append("ERROR")
    
        elif Length == 2:
            for index, word in enumerate(Word_list):
                if index == 0 and word == "vien":
                    Num_seperator = str(int(Split_num[0]) - (int(Split_num[0]) // 100) * 100)
                    if Num_seperator in Word_endings_2:
                        middle_number = Word_endings_2[Num_seperator] + " eurų"
                        Completed_list_1.append(middle_number)
                    else:
                        Completed_list_1.append("error")
    
                elif index == 0 and word == "null":
                    Completed_list_1.append("")
    
                elif index == 0 and word != "vien":
                    combined_2 = word + Word_endings_3[word]
                    if Split_num[0][1] == "0":
                        Completed_list_1.append(combined_2 + " eurų")
                    else:
                        Completed_list_1.append(combined_2)
    
                elif index == 1 and word == "null":
                    Null_val = Word_endings_1[word]
                    Null_val.split(' ')
                    Selected_value = Null_val[0]
                    Completed_list_1.append(" " + Selected_value)
    
                elif index == 1 and Split_num[0][0] == "1":
                    Completed_list_1.append("")
    
                elif index == 1 and word != 'null':
                    combined_3 = word + Word_endings_1[word]
                    if combined_3 == "vienas":
                        Completed_list_1.append(" " + combined_3 + " euras")
                    else:
                        Completed_list_1.append(" " + combined_3 + " eurai")
    
                else:
                    Completed_list_1.append("ERROR")
    
        elif Length == 1:
            for index, word in enumerate(Word_list):
                if index == 0 and word == "null":
                    Null_val = Word_endings_1[word]
                    Null_val.split(' ')
                    Selected_value = Null_val[0]
                    Completed_list_1.append("nulis eurų" + Selected_value)
    
                elif index == 0 and word != 'null':
                    combined = word + Word_endings_1[word]
                    if combined == "vienas":
                        Completed_list_1.append(combined + " euras")
                    else:
                        Completed_list_1.append(combined + " eurai")
    
                else:
                    Completed_list_1.append("ERROR")
    
        if (int(Split_num[0]) - 100) == 10:
            Completed_list_1.pop(-1)
            
        elif Length == 4 and Split_num[0][2] == "1":
            Completed_list_1.pop(-1)
            
        elif Length == 3 and Split_num[0][1] == "1":
            Completed_list_1.pop(-1)
    
        
        Cents = str(Split_num[1])
        if len(Cents) == 1:
            Completed_list_1.append(" " + Cents + "0")
        else:
            Completed_list_1.append(" " + Cents)
        
        
        Completed_list_1.insert(-1, " ir ")
        Completed_list_1.append(" ct.")
        Final_list = "".join(Completed_list_1)
        Cleaner_list = re.sub(r'\s+', ' ', Final_list)
    
        return Cleaner_list.capitalize()
        
    Word_sum = Sum_to_word_converted(final_total)
    
    Cell_40 = WS[str(Moving_cells_columns[0]) + str(18 + Actuall_cells)]
    Cell_40.value = "Suma žodžiais:"
    Cell_40.font = Font(name="Calibri", size=11, bold=True)
    Cell_40.alignment = Alignment(vertical="bottom")
    
    Cell_41 = WS[str(Moving_cells_columns[2]) + str(18 + Actuall_cells)]
    Cell_41.value = Sum_to_word_converted(final_total)
    Cell_41.font = Font(name="Calibri", size=11)
    Cell_41.alignment = Alignment(vertical="bottom")


    ## Rest moving part:
    
    Cell_42 = WS[str(Moving_cells_columns[9]) + str(18 + Actuall_cells)]
    Cell_42.value = "Iš viso (be PVM):"
    Cell_42.font = Font(name="Calibri", size=9, bold=True)
    Cell_42.alignment = Alignment(vertical="bottom")
    
    Cell_43 = WS[str(Moving_cells_columns[9]) + str(19 + Actuall_cells)]
    Cell_43.value = "PVM"
    Cell_43.font = Font(name="Calibri", size=9, bold=True)
    Cell_43.alignment = Alignment(vertical="bottom")
    
    Cell_44 = WS[str(Moving_cells_columns[11]) + str(19 + Actuall_cells)]
    Cell_44.value = "21%"
    Cell_44.font = Font(name="Calibri", size=9)
    Cell_44.alignment = Alignment(horizontal="right", vertical="bottom")
    Cell_44.border = Border(bottom=Side(style="thin", color="D9D9D9"))
    
    Cell_45 = WS[str(Moving_cells_columns[9]) + str(20 + Actuall_cells)]
    Cell_45.value = "PVM suma:"
    Cell_45.font = Font(name="Calibri", size=9, bold=True)
    Cell_45.alignment = Alignment(vertical="bottom")
    
    Cell_46 = WS[str(Moving_cells_columns[9]) + str(21 + Actuall_cells)]
    Cell_46.value = "Apmokėti suma:"
    Cell_46.font = Font(name="Calibri", size=9, bold=True)
    
    Final_value = f'={str(Moving_cells_columns[11])}{str(18 + Actuall_cells)}+{str(Moving_cells_columns[11])}{str(20 + Actuall_cells)}'
    
    Cell_47 = WS[str(Moving_cells_columns[11]) + str(21 + Actuall_cells)]
    Cell_47.value = Final_value
    Cell_47.font = Font(name="Calibri", size=9.5, bold=True)
    Cell_47.alignment = Alignment(horizontal="right", vertical="bottom")
    Cell_47.number_format = '#,##0.00'
    Cell_47.border = Border(bottom=Side(style="thin", color="D9D9D9"))
    
    Cell_48 = WS[str(Moving_cells_columns[0]) + str(21 + Actuall_cells)]
    Cell_48.value = "Sąskaitą išrašė:"
    Cell_48.font = Font(name="Calibri", size=11, bold=True)
    Cell_48.alignment = Alignment(vertical="bottom")
    
    
    D21_merge = f'{str(Moving_cells_columns[3])}{str(21 + Actuall_cells)}'
    G21_merge = f'{str(Moving_cells_columns[6])}{str(21 + Actuall_cells)}'
    Merge_DG_21 = f'{D21_merge}:{G21_merge}'
    
    WS.merge_cells(Merge_DG_21)
    
    Cell_49 = WS[str(Moving_cells_columns[3]) + str(21 + Actuall_cells)]
    Cell_49.value = "Manager Kmito"
    Cell_49.font = Font(name="Calibri", size=11, bold=True)
    Cell_49.alignment = Alignment(horizontal="center", vertical="bottom")

    
    Border_cells_1 = [(str(Moving_cells_columns[0]) + str(21 + Actuall_cells)), # A21
                      (str(Moving_cells_columns[1]) + str(21 + Actuall_cells)), # B21
                      (str(Moving_cells_columns[2]) + str(21 + Actuall_cells)), # C21
                      (str(Moving_cells_columns[3]) + str(21 + Actuall_cells)), # D21
                      (str(Moving_cells_columns[4]) + str(21 + Actuall_cells)), # E21
                      (str(Moving_cells_columns[5]) + str(21 + Actuall_cells)), # F21
                      (str(Moving_cells_columns[6]) + str(21 + Actuall_cells)), # G21
                      (str(Moving_cells_columns[7]) + str(21 + Actuall_cells))] # H21
    
    for col in Border_cells_1:
        Border_cell_col = WS[col]
        Border_cell_col.border = Border(bottom=Side(style="thin", color="000000"))
    
    
    Cell_50 = WS[str(Moving_cells_columns[0]) + str(23 + Actuall_cells)]
    Cell_50.value = "Prekes/paslaugas gavo:"
    Cell_50.font = Font(name="Calibri", size=11, bold=True)
    Cell_50.alignment = Alignment(vertical="bottom")
    
    
    F23_merge = f'{str(Moving_cells_columns[5])}{str(23 + Actuall_cells)}'
    H23_merge = f'{str(Moving_cells_columns[7])}{str(23 + Actuall_cells)}'
    Merge_FH_23 = f'{F23_merge}:{H23_merge}'
    
    WS.merge_cells(Merge_FH_23)
    Cell_51 = WS[str(Moving_cells_columns[5]) + str(23 + Actuall_cells)]
    Cell_51.value = "Pretenzijų neturiu"
    Cell_51.font = Font(name="Calibri", size=11, bold=True)
    Cell_51.alignment = Alignment(horizontal="right", vertical="bottom")
    
    
    Border_cells_2 = [(str(Moving_cells_columns[0]) + str(23 + Actuall_cells)), # A23
                      (str(Moving_cells_columns[1]) + str(23 + Actuall_cells)), # B23
                      (str(Moving_cells_columns[2]) + str(23 + Actuall_cells)), # C23
                      (str(Moving_cells_columns[3]) + str(23 + Actuall_cells)), # D23
                      (str(Moving_cells_columns[4]) + str(23 + Actuall_cells)), # E23
                      (str(Moving_cells_columns[5]) + str(23 + Actuall_cells)), # F23
                      (str(Moving_cells_columns[6]) + str(23 + Actuall_cells)), # G23
                      (str(Moving_cells_columns[7]) + str(23 + Actuall_cells))] # H23
    
    for col in Border_cells_2:
        Border_cell_col_2 = WS[col]
        Border_cell_col_2.border = Border(bottom=Side(style="thin", color="000000"))
    
    
    A24_merge = f'{str(Moving_cells_columns[0])}{str(24 + Actuall_cells)}'
    H25_merge = f'{str(Moving_cells_columns[7])}{str(25 + Actuall_cells)}'
    Merge_AH_24_25 = f'{A24_merge}:{H25_merge}'
    
    WS.merge_cells(Merge_AH_24_25)
    Cell_52 = WS[str(Moving_cells_columns[0]) + str(24 + Actuall_cells)]
    Cell_52.value = "Susipažinau su teikiamomis paslaugomis ir prekių garantijos sąlygomis, pretenzijų neturiu, sutinku su jomis"
    Cell_52.font = Font(name="Calibri", size=8)
    Cell_52.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    return WB

def Clean_filename(filename):
    return re.sub(r'[\\/:*?"<>|]', '', filename)

def File_printer(change):
    with Output:
        Output.clear_output()

        invoice_number = Invoice_name_text.value.strip()
        if not invoice_number:
            print("Please enter an invoice number.")
            return
        
        if 'Invoice' not in globals() or Invoice.empty:
            print("Please generate an invoice first")
            return

        Company_name = Clean_filename(Client.value)

        if Company_name == "Company A":
            Company_name = "Company A"

        if Company_name == "Company B":
            Company_name = "Company B"
        
        filename = f"{Company_name} Sąskaita LES {invoice_number}.xlsx"
        filepath = os.path.join(Saving_path, filename)
        
        if os.path.exists(filepath):
            print("File name already exists, please change invoice number")
            return
        
        try:
            WB = create_excel_template(Invoice, invoice_number)
            
            if Client_menu.value == Default:
                current_client_data = {
                    "Pirkėjas": Client.value,
                    "Adresas": Address.value,
                    "Įmonės kodas": Company_code.value,
                    "PVM kodas": VAT_code.value}
            else:
                current_client_data = Client_menu.value
            
            update_client_cells(current_client_data)
            WB.save(filepath)
            print(f"Invoice saved to: {filepath}")
        except Exception as e:
            print(f"An error occurred: {e}")

    
Invoice_name_text = widgets.Text(description="Invoice Number", layout=widgets.Layout(margin='0 0 0 -25px', width='300px'),\
                                 placeholder="e.g., 00001", style={'description_width': '120px'})

Invoice_text_box = widgets.HBox([Invoice_name_text, Client_menu])


#===============================================================================================================================================#
# 14. Display:

    
Header_1 = widgets.HTML(f'<h2 style="color: rgb({UI_color})">Datasets:<h2>')
Line_1 = widgets.HTML(f"<hr style='border: 1.5px solid rgb({UI_color});'>")

Show_data_button = widgets.Button(description="Show Dataset", layout=widgets.Layout(margin='20px 0 30px 5px'), button_style='primary')
Show_invoice_button = widgets.Button(description="Show invoice", layout=widgets.Layout(margin='20px 0 30px 15px'), button_style='primary')
Print_invoice_button = widgets.Button(description="Print invoice", layout=widgets.Layout(margin='20px 0 30px 15px'), button_style='primary')

Show_data_button.on_click(Invoice_filtering)
Show_invoice_button.on_click(Invoice_generator)
Print_invoice_button.on_click(File_printer)

Button_hbox = widgets.HBox([Show_data_button, Show_invoice_button, Print_invoice_button])

left_column = widgets.VBox([widgets.Label("Select categories:"), Category_box])                
right_column = widgets.VBox([widgets.Label("Select groups:"),Item_box])
layout_row = widgets.HBox([left_column, right_column])

Header_2 = widgets.HTML(f'<h2 style="; color: rgb({UI_color})">Filters:<h2>')
Line_2 = widgets.HTML(f"<hr style='border: 1.5px solid rgb({UI_color});'>")
Header_2_2 = widgets.HTML(f""" <h2 style="color: rgb({UI_color});">Special Filters:</h2> """)

Header_row = widgets.HBox([Header_2, Header_2_2], layout=widgets.Layout(justify_content='space-between',  width='100%', margin='20px 0 0 0'))

Header_3 = widgets.HTML(f'<h2 style="margin-top: 20px; color: rgb({UI_color})">Client info:<h2>')
Line_3 = widgets.HTML(f"<hr style='border: 1.5px solid rgb({UI_color});'>")

Header_4 = widgets.HTML(f'<h2 style="margin-top: 20px; color: rgb({UI_color})">Invoice settings:<h2>')
Line_4 = widgets.HTML(f"<hr style='border: 1.5px solid rgb({UI_color});'>")


display(Header_1, Line_1)
display(layout_row)
display(Header_row, Line_2)

display(assemble_filter_layout())

display(Header_3, Line_3)
display(Invoice_text_box)
display(Client_info_box)

display(Header_4, Line_4)
display(Invoice_price_settings)

display(Button_hbox)
display(Output)


#===============================================================================================================================================#
# Testing:

